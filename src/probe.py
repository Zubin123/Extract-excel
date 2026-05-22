"""
Probe a folder of timesheet workbooks and report structural variation.

For each workbook, for each non-reference sheet, find:
  - sheet count and names
  - where the day labels (MON/TUE/...) actually sit (row, columns)
  - where the date row is
  - where time rows (start/lunch/stop) are
  - where the daily totals row sits (the row with numeric pay-type hours)
  - where the grand-total cells sit
  - A1 / B2 anchor values (employee name, EE ID)
  - whether anchors match the baseline layout exactly

Outputs:
  - output/probe_report.xlsx  (per-sheet rows + per-file summary)
  - console summary of conformance

No mutation of inputs. Pure read.
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import datetime, time
from pathlib import Path

import openpyxl
import pandas as pd

REFERENCE_SHEET_HINTS = {
    "master template", "certified codes", "employees",
    "job details", "job list", "columnlists",
}

DAY_TOKENS = {"MON", "TUE", "WED", "THUR", "THU", "FRI", "SAT", "SUN"}

# Baseline layout (from the original sample) — used only for conformance reporting.
BASELINE = {
    "name_cell":      "A1",
    "id_cell":        "B2",
    "wc_state_cell":  "G13",
    "st_cell":        "H13",
    "day_row":        4,
    "date_row":       3,
    "start_row":      5,
    "lunch_out_row":  6,
    "lunch_in_row":   7,
    "stop_row":       8,
    "total_hrs_row":  9,
    "pay_totals_row": 24,
    "day_start_cols": ["L", "R", "X", "AD", "AJ", "AP", "AV"],
    "grand_total_cells_pay":   ["BB24", "BC24", "BD24", "BE24", "BF24", "BG24"],
    "grand_total_cell_hours":  "BB9",
}


def col_letter(idx: int) -> str:
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def is_reference_sheet(name: str) -> bool:
    n = name.strip().lower()
    return any(hint in n for hint in REFERENCE_SHEET_HINTS)


def is_time_value(v) -> bool:
    if isinstance(v, (datetime, time)):
        return True
    if isinstance(v, float) and 0.0 <= v < 1.0:
        return True
    return False


def scan_sheet(ws) -> dict:
    """Return a dict describing this sheet's structural landmarks."""
    max_r = min(ws.max_row or 0, 60)
    max_c = min(ws.max_column or 0, 80)

    # 1. Find the day-label row: row containing ≥3 of the DAY_TOKENS
    day_row = None
    day_cols: list[int] = []
    for r in range(1, max_r + 1):
        hits = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() in DAY_TOKENS:
                hits.append(c)
        if len(hits) >= 3:
            day_row = r
            day_cols = hits
            break

    # 2. Find a row near day_row that has ≥3 date-like values
    date_row = None
    if day_row is not None:
        for r in (day_row - 1, day_row + 1, day_row - 2, day_row + 2):
            if r < 1 or r > max_r:
                continue
            date_hits = sum(
                1 for c in day_cols if isinstance(ws.cell(row=r, column=c).value, datetime)
            )
            if date_hits >= 3:
                date_row = r
                break

    # 3. Find the daily pay-types totals row: the row where ≥4 day_cols have numeric
    #    values *and* the next 5 columns after each day_col are also mostly numeric.
    #    Scan rows below day_row.
    pay_totals_row = None
    if day_row is not None and day_cols:
        for r in range(day_row + 1, max_r + 1):
            block_numeric_hits = 0
            for c in day_cols:
                count = sum(
                    1 for k in range(6)
                    if isinstance(ws.cell(row=r, column=c + k).value, (int, float))
                )
                if count >= 4:
                    block_numeric_hits += 1
            if block_numeric_hits >= len(day_cols) - 1:
                pay_totals_row = r
                break

    # 4. Time rows: scan rows between day_row+1 and pay_totals_row-1 (or +10) for
    #    rows where ≥3 day_cols contain time-like values.
    time_rows: list[int] = []
    upper = pay_totals_row - 1 if pay_totals_row else (day_row + 12 if day_row else max_r)
    if day_row is not None:
        for r in range(day_row + 1, min(upper, max_r) + 1):
            hits = sum(1 for c in day_cols if is_time_value(ws.cell(row=r, column=c).value))
            if hits >= 3:
                time_rows.append(r)

    # 5. Total Hours daily row: row where day_cols contain numeric (not time) values
    #    between the time rows and pay_totals_row. Pick the row directly above
    #    pay_totals_row that has ≥3 numeric values across day_cols.
    total_hrs_row = None
    if day_row is not None and day_cols:
        # candidate window: after the last time row, before pay_totals_row
        lo = (time_rows[-1] + 1) if time_rows else day_row + 1
        hi = (pay_totals_row - 1) if pay_totals_row else max_r
        for r in range(lo, min(hi, max_r) + 1):
            hits = sum(
                1 for c in day_cols
                if isinstance(ws.cell(row=r, column=c).value, (int, float))
                and not is_time_value(ws.cell(row=r, column=c).value)
            )
            # string error like "#VALUE!" also counts as "the total-hours cell exists"
            err_hits = sum(
                1 for c in day_cols
                if isinstance(ws.cell(row=r, column=c).value, str)
                and "#" in (ws.cell(row=r, column=c).value or "")
            )
            if hits + err_hits >= 3:
                total_hrs_row = r
                break

    # 6. Grand-total columns (BB..BG in baseline): on pay_totals_row, find the last
    #    contiguous run of 6 numeric cells to the right of the last day block.
    grand_total_cols: list[int] = []
    if pay_totals_row is not None and day_cols:
        last_block_end = day_cols[-1] + 5  # 6 columns per block
        # Scan rightward for the next group of ≥6 numeric-or-zero cells
        for c in range(last_block_end + 1, max_c + 1):
            run = []
            for k in range(8):
                v = ws.cell(row=pay_totals_row, column=c + k).value
                if isinstance(v, (int, float)) or v is None:
                    run.append(c + k)
                else:
                    break
            if len(run) >= 6:
                grand_total_cols = run[:6]
                break

    # 7. Anchor cells
    a1 = ws["A1"].value
    b2 = ws["B2"].value
    g13 = ws["G13"].value
    h13 = ws["H13"].value

    return {
        "sheet":            ws.title,
        "max_row":          ws.max_row,
        "max_col":          ws.max_column,
        "A1":               repr(a1)[:60] if a1 is not None else "",
        "B2":               repr(b2)[:30] if b2 is not None else "",
        "G13":              repr(g13)[:30] if g13 is not None else "",
        "H13":              repr(h13)[:30] if h13 is not None else "",
        "day_row":          day_row,
        "day_cols":         ",".join(col_letter(c) for c in day_cols),
        "n_day_cols":       len(day_cols),
        "date_row":         date_row,
        "time_rows":        ",".join(str(r) for r in time_rows),
        "total_hrs_row":    total_hrs_row,
        "pay_totals_row":   pay_totals_row,
        "grand_total_cols": ",".join(col_letter(c) for c in grand_total_cols),
    }


def conforms_to_baseline(sig: dict) -> tuple[bool, list[str]]:
    deltas: list[str] = []
    if sig["day_row"] != BASELINE["day_row"]:
        deltas.append(f"day_row={sig['day_row']} (expected {BASELINE['day_row']})")
    if sig["date_row"] != BASELINE["date_row"]:
        deltas.append(f"date_row={sig['date_row']} (expected {BASELINE['date_row']})")
    if sig["total_hrs_row"] != BASELINE["total_hrs_row"]:
        deltas.append(f"total_hrs_row={sig['total_hrs_row']} (expected {BASELINE['total_hrs_row']})")
    if sig["pay_totals_row"] != BASELINE["pay_totals_row"]:
        deltas.append(f"pay_totals_row={sig['pay_totals_row']} (expected {BASELINE['pay_totals_row']})")
    if sig["day_cols"] != ",".join(BASELINE["day_start_cols"]):
        deltas.append(f"day_cols={sig['day_cols']} (expected {','.join(BASELINE['day_start_cols'])})")
    expected_gt = ",".join(c.rstrip("0123456789") for c in BASELINE["grand_total_cells_pay"])
    if sig["grand_total_cols"] != expected_gt:
        deltas.append(f"grand_total_cols={sig['grand_total_cols']} (expected {expected_gt})")
    return (len(deltas) == 0), deltas


def probe_workbook(path: Path, batch: str = "") -> tuple[list[dict], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    sheets_info: list[dict] = []
    employee_sheets = 0
    reference_sheets = 0
    conforming = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if is_reference_sheet(name):
            reference_sheets += 1
            sheets_info.append({
                "batch": batch,
                "file": path.name,
                "sheet": name,
                "kind": "reference",
                "conforms": "",
                "deltas": "",
            })
            continue

        sig = scan_sheet(ws)
        employee_sheets += 1
        conforms, deltas = conforms_to_baseline(sig)
        if conforms:
            conforming += 1
        sheets_info.append({
            "batch":  batch,
            "file":   path.name,
            "kind":   "employee",
            "conforms": "YES" if conforms else "NO",
            "deltas": "; ".join(deltas),
            **sig,
        })

    file_summary = {
        "batch":            batch,
        "file":             path.name,
        "total_sheets":     len(wb.sheetnames),
        "reference_sheets": reference_sheets,
        "employee_sheets":  employee_sheets,
        "conforming":       conforming,
        "non_conforming":   employee_sheets - conforming,
        "all_match":        "YES" if conforming == employee_sheets and employee_sheets > 0 else "NO",
    }
    return sheets_info, [file_summary]


def main():
    root = Path(__file__).parent.parent
    tagged: list[tuple[Path, str]] = []
    baseline_file = root / "data" / "WE 01 08 22 CA GF'S - WA 4939-0899-6263_1.xlsx"
    if baseline_file.exists():
        tagged.append((baseline_file, "baseline"))
    for sub in ("sample", "sample1"):
        sample_dir = root / "data" / sub
        if sample_dir.exists():
            for f in sorted(sample_dir.glob("*.xlsx")):
                tagged.append((f, sub))

    tagged = [(f, b) for (f, b) in tagged if not f.name.startswith("~$")]

    print(f"Probing {len(tagged)} workbook(s)...\n")
    all_sheets: list[dict] = []
    all_files:  list[dict] = []

    for f, batch in tagged:
        try:
            sheets, summary = probe_workbook(f, batch=batch)
            all_sheets.extend(sheets)
            all_files.extend(summary)
            s = summary[0]
            print(f"  [{batch:<8}] {f.name[:55]:<55} "
                  f"sheets={s['total_sheets']:>2} "
                  f"emp={s['employee_sheets']:>2} "
                  f"ref={s['reference_sheets']:>2} "
                  f"conform={s['conforming']:>2}/{s['employee_sheets']:<2} "
                  f"{'OK' if s['all_match']=='YES' else 'DIFF'}")
        except Exception as e:
            print(f"  {f.name}: ERROR {e!r}")
            all_files.append({"file": f.name, "total_sheets": "ERROR", "error": repr(e)})

    out = root / "output" / "probe_report.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    sheets_df = pd.DataFrame(all_sheets)
    files_df  = pd.DataFrame(all_files)

    # --- aggregate variation patterns ---
    emp_only = sheets_df[sheets_df.get("kind") == "employee"] if "kind" in sheets_df.columns else sheets_df
    patterns_rows = []
    for col in ("day_row", "date_row", "total_hrs_row", "pay_totals_row", "day_cols", "grand_total_cols", "n_day_cols"):
        if col not in emp_only.columns:
            continue
        c = Counter(emp_only[col].dropna().tolist())
        for val, n in c.most_common():
            patterns_rows.append({"field": col, "value": val, "count": n})
    patterns_df = pd.DataFrame(patterns_rows)

    with pd.ExcelWriter(out, engine="openpyxl") as w:
        files_df.to_excel(w,    sheet_name="File_Summary", index=False)
        sheets_df.to_excel(w,   sheet_name="Sheet_Details", index=False)
        patterns_df.to_excel(w, sheet_name="Variation_Patterns", index=False)

    print(f"\nReport: {out}")

    # --- console summary of variation ---
    print("\n=== Variation patterns across employee sheets ===")
    for col in ("day_row", "date_row", "total_hrs_row", "pay_totals_row", "day_cols", "grand_total_cols", "n_day_cols"):
        if col not in emp_only.columns:
            continue
        c = Counter(emp_only[col].dropna().tolist())
        top = c.most_common(5)
        print(f"  {col:<18} {len(c)} distinct value(s) — top: {top}")


if __name__ == "__main__":
    main()
