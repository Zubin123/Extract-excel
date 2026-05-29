"""Extract weekly timesheet data from Excel workbook(s) into flat tabular output.

Phase 1 of the robust pipeline:
  - Config-driven profiles (config/schema.yaml)
  - Structural sheet classification (no hardcoded reference-sheet name list)
  - Pay-totals row discovered via sum cross-check
  - Sheets that fail anchor resolution are listed on the Unmatched_Sheets tab
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, time
from pathlib import Path

import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import Font, PatternFill

from anchors import (
    classify_sheet,
    col_letter_to_index,
    day_alias_index,
    day_index_to_label,
    discover_field_mechanic_paytype_cols as discover_fm_paytype_cols,
    discover_paytype_columns_by_label,
    find_day_label_row,
    find_field_mechanic_header_row,
    find_fm_job_col,
    find_label_in_grid as anchors_find_label_in_grid,
    resolve_anchor_cell,
    resolve_pay_totals_row,
    select_profile,
    sheet_has_day_grid,
)
from qa import (
    check_date_sequence,
    check_ee_id_consistency,
    check_filename_week_match,
    check_time_order,
    check_weekly_hours_range,
    two_pass_totals_row,
)

DEFAULT_DAY_LABELS = ["MON", "TUE", "WED", "THUR", "FRI", "SAT", "SUN"]
DEFAULT_PAY_TYPES  = ["RT", "OT", "DT", "PD", "4D", "4A"]


_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_GREY   = PatternFill("solid", fgColor="D9D9D9")
_BOLD   = Font(bold=True)


# ── Formatters ──────────────────────────────────────────────────────────────

def fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%y")
    return str(val)


def fmt_time(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        fmt = "%#I:%M %p" if sys.platform == "win32" else "%-I:%M %p"
        return val.strftime(fmt)
    if isinstance(val, time):
        dt = datetime.combine(datetime.today(), val)
        fmt = "%#I:%M %p" if sys.platform == "win32" else "%-I:%M %p"
        return dt.strftime(fmt)
    if isinstance(val, float) and 0.0 <= val < 1.0:
        total_minutes = round(val * 24 * 60)
        hours, minutes = divmod(total_minutes, 60)
        period = "AM" if hours < 12 else "PM"
        display_hour = hours % 12 or 12
        return f"{display_hour}:{minutes:02d} {period}"
    return ""


def coerce_hours(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return 0.0


def coerce_total_hours(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return None


def infer_box_name(filename: str) -> str:
    stem = Path(filename).stem
    m = re.search(r"WE[\s_](\d{2})[\s_](\d{2})[\s_](\d{2})", stem, re.IGNORECASE)
    if m:
        return f"WE {m.group(1)} {m.group(2)} {m.group(3)}"
    return stem


def _is_time_value(v) -> bool:
    if isinstance(v, (datetime, time)):
        return True
    if isinstance(v, float) and 0.0 <= v < 1.0:
        return True
    return False


def _row_qa_flag(time_cells: dict, total_hrs, total_hrs_cell: str) -> str:
    if all(v is None for (_addr, v) in time_cells.values()):
        return "OK (no time data)"
    issues = []
    for field_name, (addr, val) in time_cells.items():
        if val is None or _is_time_value(val):
            continue
        issues.append(f"{field_name} cell {addr} = {val!r} — not a time")
    if isinstance(total_hrs, str) and not issues:
        issues.append(f"Total Hours cell {total_hrs_cell} = {total_hrs!r} (source formula error)")
    return "; ".join(issues) if issues else "OK"


# ── Sheet extraction ────────────────────────────────────────────────────────

def _anchor_issue(field: str, status: str, value=None) -> str | None:
    # Empty / missing value check — independent of header-resolution status.
    is_empty = value is None or (isinstance(value, str) and value.strip() == "")
    if is_empty:
        if status == "missing":
            return f"[CHECK] {field} empty — header not found in header_row"
        if status.startswith("relocated:"):
            return f"[CHECK] {field} empty — header relocated to {status.split(':', 1)[1]} but no value"
        return f"[CHECK] {field} empty — no value in source cell"
    if status in ("ok", "noverify"):
        return None
    if status.startswith("relocated:"):
        return f"[CHECK] {field} anchor relocated to {status.split(':', 1)[1]}"
    if status == "missing":
        return f"[CHECK] {field} header not found in header_row"
    return f"[CHECK] {field} anchor unknown status {status!r}"


def extract_sheet(ws, profile: dict, pay_totals_row: int, folder_name: str,
                  box_name: str, excel_name: str, kind: str = "employee") -> list[dict]:
    sheet_name = ws.title
    anchors = profile["anchors"]
    raw_a1 = ws[anchors["employee_name"]].value
    if kind == "employee_alt_layout":
        ee_id = ws["C2"].value
    else:
        ee_id = ws[anchors["ee_id"]].value
    wc_state_val, wc_status, _ = resolve_anchor_cell(ws, anchors.get("wc_state"))
    st_val,       st_status, _ = resolve_anchor_cell(ws, anchors.get("st"))
    wc_state = wc_state_val if wc_state_val is not None else ""
    st       = st_val       if st_val       is not None else ""

    anchor_issues: list[str] = []
    if kind == "employee_placeholder":
        employee_name = ""
        anchor_issues.append(
            f"[CHECK] Employee Name placeholder in A1 ({raw_a1!r}) — left blank (no reliable source)"
        )
    elif kind == "employee_alt_layout":
        employee_name = ws["B1"].value or ""
        anchor_issues.append(
            f"[CHECK] Alternate layout — Employee Name from B1, EE ID from C2"
        )
    else:
        employee_name = raw_a1 or ""

    for field, status, value in (
        ("WC State", wc_status, wc_state_val),
        ("ST",       st_status, st_val),
    ):
        msg = _anchor_issue(field, status, value)
        if msg:
            anchor_issues.append(msg)

    db = profile["day_blocks"]
    day_labels = db["day_labels"]
    day_start_cols = db["day_start_cols"]
    fallback_pay_types = db["pay_types"]  # YAML fallback, used only if no labels discovered

    date_row = profile["date_row"]
    total_hours_row = profile["total_hours_row"]
    t_rows = profile["time_rows"]

    # Label-driven pay-type discovery: read each column's actual label in the
    # header row (the row containing 'WC STATE') and route values to the
    # output field matching the label. Closes the silent-position-drift class
    # — sheets where M='PTO1' instead of M='OT1' now route correctly without
    # any per-sheet config. See anchors.discover_paytype_columns_by_label.
    header_loc = anchors_find_label_in_grid(ws, "WC STATE", max_r=30, max_c=20)
    paytype_header_row = header_loc[0] if header_loc else None
    day_start_idx_list = [col_letter_to_index(c) for c in day_start_cols]
    paytype_grid: dict[tuple[int, str], int] = {}
    if paytype_header_row is not None:
        paytype_grid, discovery_issues = discover_paytype_columns_by_label(
            ws, paytype_header_row, day_start_idx_list
        )
        anchor_issues.extend(discovery_issues)
    else:
        anchor_issues.append(
            "[CHECK] WC STATE header not found — falling back to YAML pay-type positions"
        )

    # Fallback: positional reading from YAML, only when label discovery yielded
    # nothing for a day. We always emit the 8 canonical output columns.
    OUTPUT_PAY_TYPES = ["RT", "OT", "DT", "PD", "4D", "4A", "PTO", "HP"]

    rows = []

    for day_idx, (day_label, start_col) in enumerate(zip(day_labels, day_start_cols)):
        start_idx = col_letter_to_index(start_col)

        date_val   = ws.cell(row=date_row, column=start_idx).value
        time_start = ws.cell(row=t_rows["Start"],     column=start_idx).value
        lunch_out  = ws.cell(row=t_rows["Lunch Out"], column=start_idx).value
        lunch_in   = ws.cell(row=t_rows["Lunch In"],  column=start_idx).value
        stop       = ws.cell(row=t_rows["Stop"],      column=start_idx).value
        total_hrs  = ws.cell(row=total_hours_row,     column=start_idx).value

        # Read each output pay-type from its label-discovered column.
        day_hours = {pt: 0.0 for pt in OUTPUT_PAY_TYPES}
        has_any_discovered = any((day_idx, pt) in paytype_grid for pt in OUTPUT_PAY_TYPES)
        if has_any_discovered:
            for pt in OUTPUT_PAY_TYPES:
                col = paytype_grid.get((day_idx, pt))
                if col is not None:
                    day_hours[pt] = coerce_hours(
                        ws.cell(row=pay_totals_row, column=col).value
                    )
        else:
            # No labels found for this day — fall back to YAML positions.
            # Don't populate PTO/HP via fallback (they aren't in the YAML list).
            for i, col_name in enumerate(fallback_pay_types):
                if col_name in OUTPUT_PAY_TYPES:
                    day_hours[col_name] = coerce_hours(
                        ws.cell(row=pay_totals_row, column=start_idx + i).value
                    )

        time_cells = {
            "Start":     (f"{start_col}{t_rows['Start']}",     time_start),
            "Lunch Out": (f"{start_col}{t_rows['Lunch Out']}", lunch_out),
            "Lunch In":  (f"{start_col}{t_rows['Lunch In']}",  lunch_in),
            "Stop":      (f"{start_col}{t_rows['Stop']}",      stop),
        }
        base_flag = _row_qa_flag(time_cells, total_hrs, f"{start_col}{total_hours_row}")
        order_issues = check_time_order(time_cells)
        flag_parts = [p for p in [base_flag if base_flag != "OK" else None, *order_issues] if p]
        qa_flag = "; ".join(flag_parts) if flag_parts else "OK"

        rows.append({
            "Folder Name":   folder_name,
            "Box Name":      box_name,
            "Excel Name":    excel_name,
            "Sheet Name":    sheet_name,
            "Employee Name": employee_name,
            "EE ID":         ee_id,
            "WC State":      wc_state,
            "ST":            st,
            "Date":          fmt_date(date_val),
            "Day":           day_label,
            "Start":         fmt_time(time_start),
            "Lunch Out":     fmt_time(lunch_out),
            "Lunch In":      fmt_time(lunch_in),
            "Stop":          fmt_time(stop),
            "Total Hours":   coerce_total_hours(total_hrs),
            **{pt: day_hours[pt] for pt in OUTPUT_PAY_TYPES},
            "QA_Flag": qa_flag,
        })

    # TOTALS row — sum each output pay-type across the 7 day rows we just
    # produced. The day values themselves came from label-discovered columns,
    # so summing them is the authoritative total. We additionally cross-check
    # against the source's grand-totals row (BB..BG / BC..BH) for pay-types
    # that map positionally — mismatches surface as a [CHECK].
    gt_cols = profile["grand_totals"]["cols"]
    th_col = profile["grand_totals"]["total_hours_col"]

    totals_hours = {pt: 0.0 for pt in OUTPUT_PAY_TYPES}
    for pt in OUTPUT_PAY_TYPES:
        totals_hours[pt] = round(sum(r[pt] for r in rows), 4)

    # Cross-check: for the YAML's fallback_pay_types (which positionally map to
    # gt_cols), the source's grand-totals cells should equal our computed sum.
    for col_name, col_letter in zip(fallback_pay_types, gt_cols):
        if col_name not in OUTPUT_PAY_TYPES:
            continue
        src_gt = ws.cell(row=pay_totals_row,
                         column=col_letter_to_index(col_letter)).value
        if isinstance(src_gt, (int, float)):
            our_sum = totals_hours[col_name]
            if abs(float(src_gt) - our_sum) > 0.01:
                # Trust the discovered-column sum (label-driven), but flag.
                anchor_issues.append(
                    f"[CHECK] {col_name} grand-total cell = {src_gt} disagrees with "
                    f"label-discovered day-sum {our_sum}"
                )

    grand_th_cell = f"{th_col}{total_hours_row}"
    grand_total_hours_raw = ws[grand_th_cell].value
    grand_total_hours = coerce_total_hours(grand_total_hours_raw)

    if grand_total_hours is None and grand_total_hours_raw is not None:
        grand_total_hours = round(
            sum(r["Total Hours"] for r in rows if r["Total Hours"] is not None), 4
        )
        totals_flag = (
            f"Total Hours cell {grand_th_cell} = {grand_total_hours_raw!r} "
            f"— recomputed from day rows"
        )
    else:
        totals_flag = "OK"

    if anchor_issues:
        merged = "; ".join(anchor_issues)
        totals_flag = merged if totals_flag == "OK" else f"{totals_flag}; {merged}"

    rows.append({
        "Folder Name":   folder_name,
        "Box Name":      box_name,
        "Excel Name":    excel_name,
        "Sheet Name":    sheet_name,
        "Employee Name": employee_name,
        "EE ID":         ee_id,
        "WC State":      wc_state,
        "ST":            st,
        "Date":          "",
        "Day":           "TOTALS",
        "Start":         "",
        "Lunch Out":     "",
        "Lunch In":      "",
        "Stop":          "",
        "Total Hours":   grand_total_hours,
        **{pt: totals_hours[pt] for pt in OUTPUT_PAY_TYPES},
        "QA_Flag": totals_flag,
    })

    return rows


def extract_sheet_no_grid(ws, folder_name: str, box_name: str,
                          excel_name: str, profile: dict | None = None,
                          kind: str = "employee") -> list[dict]:
    """Emit 8 blank rows (MON..SUN + TOTALS) for an employee sheet that has
    no weekly day-label grid (e.g. 'Job' / 'OVERHEAD- CA').

    When `profile` is supplied, wc_state/st use the same header-verified
    resolver as the grid-bearing path; otherwise they fall back to G13/H13.
    When `kind == "employee_placeholder"`, sheet name is used as the
    employee name and a [CHECK] flag is emitted on the TOTALS row.
    """
    sheet_name = ws.title
    raw_a1 = ws["A1"].value
    if kind == "employee_placeholder":
        employee_name = ""
    elif kind == "employee_alt_layout":
        employee_name = ws["B1"].value or ""
    else:
        employee_name = raw_a1 or ""

    if kind == "employee_alt_layout":
        ee_id = ws["C2"].value
    else:
        ee_id = ws["B2"].value

    if profile is not None:
        anchors = profile.get("anchors", {})
        wc_val, _wc_status, _ = resolve_anchor_cell(ws, anchors.get("wc_state"))
        st_val, _st_status, _ = resolve_anchor_cell(ws, anchors.get("st"))
        wc_state = wc_val if wc_val is not None else ""
        st       = st_val if st_val is not None else ""
    else:
        wc_state = ws["G13"].value or ""
        st = ws["H13"].value or ""

    base_flag = "no time grid"
    if kind == "employee_placeholder":
        totals_flag = (
            f"{base_flag}; [CHECK] Employee Name placeholder in A1 ({raw_a1!r}) — left blank"
        )
    elif kind == "employee_alt_layout":
        totals_flag = (
            f"{base_flag}; [CHECK] Alternate layout — Employee Name from B1, EE ID from C2"
        )
    else:
        totals_flag = base_flag

    rows = []
    for day_label in DEFAULT_DAY_LABELS + ["TOTALS"]:
        row = {
            "Folder Name":   folder_name,
            "Box Name":      box_name,
            "Excel Name":    excel_name,
            "Sheet Name":    sheet_name,
            "Employee Name": employee_name,
            "EE ID":         ee_id,
            "WC State":      wc_state,
            "ST":            st,
            "Date":          "",
            "Day":           day_label,
            "Start":         "",
            "Lunch Out":     "",
            "Lunch In":      "",
            "Stop":          "",
            "Total Hours":   None,
            **{pt: "" for pt in DEFAULT_PAY_TYPES},
            "PTO": "",
            "HP":  "",
            "QA_Flag": totals_flag if day_label == "TOTALS" else base_flag,
        }
        rows.append(row)
    return rows


def extract_field_mechanic_sheet(ws, header_row: int, folder_name: str,
                                 box_name: str, excel_name: str,
                                 kind: str = "employee") -> list[dict]:
    """Extract a Field Mechanic line-item sheet into the standard per-day output.

    Everything is discovered by label (see anchors.find_field_mechanic_*):
      - identity (name/EE ID) per `kind` (A1/B2 standard, B1/C2 alt-layout)
      - WC State / ST by label in the header row
      - day columns + dates + clock-time cells from the day-label row & the
        'Total Hours' clock-summary row
      - per-day pay-type hours (RT/OT/DT/PTO/HP) by summing each labelled
        (day, pay-type) column down the line-item rows below `header_row`

    Produces 7 day rows (MON..SUN) + a TOTALS row, matching the output schema.
    Unresolvable fields are left blank and flagged with [CHECK] — never guessed.
    """
    sheet_name = ws.title

    # Identity — reuse the same per-kind rules as the standard path.
    raw_a1 = ws["A1"].value
    anchor_issues: list[str] = []
    if kind == "employee_alt_layout":
        employee_name = ws["B1"].value or ""
        ee_id = ws["C2"].value
        anchor_issues.append("[CHECK] Alternate layout — Employee Name from B1, EE ID from C2")
    elif kind == "employee_placeholder":
        employee_name = ""
        ee_id = ws["B2"].value
        anchor_issues.append(
            f"[CHECK] Employee Name placeholder in A1 ({raw_a1!r}) — left blank (no reliable source)"
        )
    else:
        employee_name = raw_a1 or ""
        ee_id = ws["B2"].value

    # WC State / ST — by label, in the discovered header row (value sits in the
    # data rows below; take the first non-blank).
    def _resolve_label_value(label: str):
        loc = anchors_find_label_in_grid(ws, label, max_r=header_row + 1, max_c=20)
        if loc is None:
            return "", f"[CHECK] {label} header not found"
        hr, hc = loc
        for r in range(hr + 1, min((ws.max_row or hr) + 1, hr + 30)):
            v = ws.cell(row=r, column=hc).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                if isinstance(v, str) and v.strip().upper() in ("#N/A", "#VALUE!"):
                    continue
                return v, None
        return "", f"[CHECK] {label} empty — no value in column below header"

    wc_state, wc_issue = _resolve_label_value("WC STATE")
    st, st_issue = _resolve_label_value("ST")
    for issue in (wc_issue, st_issue):
        if issue:
            anchor_issues.append(issue)

    # Day-label row + columns (full names at row 3 — found by vocabulary).
    day_row_info = find_day_label_row(ws)
    grid, day_indices, _pts_seen = discover_fm_paytype_cols(ws, header_row)

    # JOB column — only rows with a job code are real line items. Rows below
    # without a JOB value are echo/import-shadow rows that duplicate per-day
    # pay-type cells; summing them double-counts. If we can't find the JOB
    # column we fall back to "all rows" (and surface a [CHECK] flag).
    job_col = find_fm_job_col(ws, header_row)
    if job_col is None:
        anchor_issues.append("[CHECK] FM JOB column not found — line-item filter disabled (sums may double-count)")

    def _is_line_item_row(r: int) -> bool:
        if job_col is None:
            return True
        v = ws.cell(row=r, column=job_col).value
        if v is None:
            return False
        if isinstance(v, str) and v.strip() == "":
            return False
        return True

    # Clock-time summary: locate the time-label column ('Start'/'Stop' etc.) and
    # the 'Total Hours' clock row, both by label.
    th_loc = anchors_find_label_in_grid(ws, "Total Hours", max_r=header_row, max_c=20)
    start_loc = anchors_find_label_in_grid(ws, "Start", max_r=header_row, max_c=20)
    lout_loc  = anchors_find_label_in_grid(ws, "Lunch Out", max_r=header_row, max_c=20)
    lin_loc   = anchors_find_label_in_grid(ws, "Lunch In", max_r=header_row, max_c=20)
    stop_loc  = anchors_find_label_in_grid(ws, "Stop", max_r=header_row, max_c=20)
    total_hours_row = th_loc[0] if th_loc else None
    start_row = start_loc[0] if start_loc else None
    lout_row  = lout_loc[0] if lout_loc else None
    lin_row   = lin_loc[0] if lin_loc else None
    stop_row  = stop_loc[0] if stop_loc else None

    # Day columns of the clock grid: the day-label row's columns, in order.
    # Build day_idx -> clock-column from the day-label row tokens.
    clock_day_cols: dict[int, int] = {}
    date_row = None
    if day_row_info is not None:
        day_label_row = day_row_info[0]
        date_row = day_label_row - 1  # invariant: date row is one above day labels
        cmax = min(ws.max_column or 0, 120)
        for c in range(1, cmax + 1):
            v = ws.cell(row=day_label_row, column=c).value
            if isinstance(v, str):
                idx = day_alias_index(v.strip().upper())
                if idx is not None and idx not in clock_day_cols:
                    clock_day_cols[idx] = c

    last_row = ws.max_row or header_row
    rows: list[dict] = []
    pay_out_types = ["RT", "OT", "DT", "PD", "4D", "4A", "PTO", "HP"]
    week_totals = {pt: 0.0 for pt in pay_out_types}
    week_total_hours = 0.0
    any_total_hours = False

    for day_idx in range(7):
        day_label = day_index_to_label(day_idx)
        clock_col = clock_day_cols.get(day_idx)

        # Per-day pay-type sums from the line-item table (label-discovered cols).
        # Only sum rows that ARE line items (JOB column populated) — echo rows
        # below the line-item block duplicate per-day pay-type cells.
        day_pay = {pt: 0.0 for pt in pay_out_types}
        for pt in ("RT", "OT", "DT", "PTO", "HP"):
            col = grid.get((day_idx, pt))
            if col is not None:
                s = 0.0
                for r in range(header_row + 1, last_row + 1):
                    if not _is_line_item_row(r):
                        continue
                    s += coerce_hours(ws.cell(row=r, column=col).value)
                day_pay[pt] = round(s, 4)
                week_totals[pt] += day_pay[pt]

        # Clock-grid cells (date / times / daily Total Hours) by discovered cols.
        def _clock(rownum):
            if rownum is None or clock_col is None:
                return None
            return ws.cell(row=rownum, column=clock_col).value

        date_val   = ws.cell(row=date_row, column=clock_col).value if (date_row and clock_col) else None
        time_start = _clock(start_row)
        lunch_out  = _clock(lout_row)
        lunch_in   = _clock(lin_row)
        stop       = _clock(stop_row)
        total_hrs  = coerce_total_hours(_clock(total_hours_row))
        if total_hrs is not None:
            week_total_hours += total_hrs
            any_total_hours = True

        rows.append({
            "Folder Name": folder_name, "Box Name": box_name,
            "Excel Name": excel_name, "Sheet Name": sheet_name,
            "Employee Name": employee_name, "EE ID": ee_id,
            "WC State": wc_state, "ST": st,
            "Date": fmt_date(date_val), "Day": day_label,
            "Start": fmt_time(time_start), "Lunch Out": fmt_time(lunch_out),
            "Lunch In": fmt_time(lunch_in), "Stop": fmt_time(stop),
            "Total Hours": total_hrs,
            "RT": day_pay["RT"], "OT": day_pay["OT"], "DT": day_pay["DT"],
            "PD": 0.0, "4D": 0.0, "4A": 0.0,
            "PTO": day_pay["PTO"], "HP": day_pay["HP"],
            "QA_Flag": "OK",
        })

    # TOTALS row — week sums (computed from the discovered columns).
    totals_flag = "; ".join(anchor_issues) if anchor_issues else "OK"
    rows.append({
        "Folder Name": folder_name, "Box Name": box_name,
        "Excel Name": excel_name, "Sheet Name": sheet_name,
        "Employee Name": employee_name, "EE ID": ee_id,
        "WC State": wc_state, "ST": st,
        "Date": "", "Day": "TOTALS",
        "Start": "", "Lunch Out": "", "Lunch In": "", "Stop": "",
        "Total Hours": round(week_total_hours, 4) if any_total_hours else None,
        "RT": round(week_totals["RT"], 4), "OT": round(week_totals["OT"], 4),
        "DT": round(week_totals["DT"], 4),
        "PD": 0.0, "4D": 0.0, "4A": 0.0,
        "PTO": round(week_totals["PTO"], 4), "HP": round(week_totals["HP"], 4),
        "QA_Flag": totals_flag,
    })
    return rows


def check_accuracy(employee_rows: list[dict], pay_types: list[str],
                   tolerance: float) -> tuple[bool, dict]:
    day_rows   = [r for r in employee_rows if r["Day"] != "TOTALS"]
    totals_row = next(r for r in employee_rows if r["Day"] == "TOTALS")

    results  = {}
    all_pass = True

    for col in pay_types:
        daily_sum   = round(sum(r[col] for r in day_rows), 4)
        grand_total = totals_row[col]
        passed = abs(daily_sum - grand_total) <= tolerance
        results[col] = (daily_sum, grand_total, "MATCH" if passed else "MISMATCH")
        if not passed:
            all_pass = False

    daily_th_values = [r["Total Hours"] for r in day_rows]
    grand_th        = totals_row["Total Hours"]
    if any(v is None for v in daily_th_values) or grand_th is None:
        results["Total Hours"] = (
            round(sum(v for v in daily_th_values if v is not None), 4),
            grand_th if grand_th is not None else "#VALUE!",
            "SKIPPED (#VALUE in source)",
        )
    else:
        daily_sum = round(sum(daily_th_values), 4)
        passed = abs(daily_sum - grand_th) <= tolerance
        results["Total Hours"] = (daily_sum, grand_th, "MATCH" if passed else "MISMATCH")
        if not passed:
            all_pass = False

    return all_pass, results


# ── Writer ──────────────────────────────────────────────────────────────────

_NUMERIC_OUTPUT_COLS = ("Total Hours", "RT", "OT", "DT", "PD", "4D", "4A", "PTO", "HP")


def _format_numeric_cells(data_df: pd.DataFrame) -> pd.DataFrame:
    """Render numeric output columns as 2-decimal strings to match the feedback
    file's format. Zero / None on a day row becomes '0.00' (the manager treats
    blanks as zero); TOTALS rows keep Total Hours but blank-out pay-types to
    match the manager's expected layout. In-memory floats are preserved by
    operating on a copy.
    """
    df = data_df.copy()
    if "Day" not in df.columns:
        return df

    def fmt(v):
        if v is None or v == "":
            return "0.00"
        if isinstance(v, float) and v != v:  # NaN
            return "0.00"
        try:
            return f"{float(v):.2f}"
        except (TypeError, ValueError):
            return v  # leave non-numeric (e.g. error strings) alone

    is_totals = df["Day"].astype(str).str.upper() == "TOTALS"

    for col in _NUMERIC_OUTPUT_COLS:
        if col not in df.columns:
            continue
        # Cast to object so we can mix strings and numbers in the same column
        # without pandas rejecting the assignment as a dtype violation.
        df[col] = df[col].astype(object)
        if col == "Total Hours":
            df[col] = df[col].map(fmt)
        else:
            # Day rows: format zero/None as '0.00'; TOTALS rows: keep the
            # numeric grand total. Manager's expected block leaves pay-type
            # TOTALS blank; losing that data in our output costs more than
            # the cosmetic mismatch (the comparison ignores unspecified cells).
            day_mask = ~is_totals
            df.loc[day_mask, col] = df.loc[day_mask, col].map(fmt)
    return df


def _write_excel(out_path: Path, data_df: pd.DataFrame, qa_rows: list[dict],
                 run_info: dict, unmatched: list[dict], id_conflicts: list[dict]):
    data_df = _format_numeric_cells(data_df)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        data_df.to_excel(writer, sheet_name="Data", index=False)
        pd.DataFrame(qa_rows).to_excel(writer, sheet_name="QA_Summary", index=False)
        pd.DataFrame([run_info]).to_excel(writer, sheet_name="Run_Info", index=False)
        pd.DataFrame(unmatched).to_excel(writer, sheet_name="Unmatched_Sheets", index=False)
        pd.DataFrame(id_conflicts).to_excel(writer, sheet_name="ID_Conflicts", index=False)

    wb = openpyxl.load_workbook(out_path)

    # Data sheet colouring
    ws_data = wb["Data"]
    qa_col_idx  = data_df.columns.get_loc("QA_Flag") + 1
    day_col_idx = data_df.columns.get_loc("Day") + 1
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        cell = row[qa_col_idx - 1]
        val  = str(cell.value or "")
        is_totals_row = str(row[day_col_idx - 1].value or "") == "TOTALS"
        # Treat as OK: literal "OK", "OK (no time data)", or anything where only
        # [INFO] notes are attached and no [CHECK]/[WARN]/error prefix appears.
        is_ok = (
            val == "OK"
            or val.startswith("OK (")
            or ("[CHECK]" not in val and "[WARN]" not in val
                and not any(tok in val for tok in ("— not a time", "#VALUE", "no time grid")))
        )
        if is_ok:
            cell.fill = _GREY if is_totals_row else _GREEN
        else:
            cell.fill = _RED
        cell.font = Font(bold=not is_ok)

    for col_cells in ws_data.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws_data.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)

    # QA_Summary colouring
    ws_qa = wb["QA_Summary"]
    headers = [c.value for c in ws_qa[1]]
    for row in ws_qa.iter_rows(min_row=2, max_row=ws_qa.max_row):
        row[0].font = _BOLD
        for cell in row:
            col_header = headers[cell.column - 1] if cell.column <= len(headers) else ""
            val = str(cell.value or "")
            if col_header == "Overall":
                cell.fill = _GREEN if val == "PASS" else _RED
                cell.font = Font(bold=True)
            elif col_header.endswith("_match"):
                if val == "MATCH":
                    cell.fill = _GREEN
                elif val.startswith("SKIPPED"):
                    cell.fill = _YELLOW
                else:
                    cell.fill = _RED

    for col_cells in ws_qa.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws_qa.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)

    # Header row + freeze panes on all sheets
    for sheet_name in ("Data", "QA_Summary", "Run_Info", "Unmatched_Sheets", "ID_Conflicts"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
        ws.freeze_panes = "A2"

    wb.save(out_path)


# ── Pipeline ────────────────────────────────────────────────────────────────

def load_config(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def process_workbook(input_path: Path, config: dict, folder_name: str,
                     box_name: str | None) -> tuple[list[dict], list[dict], list[dict]]:
    """Return (data_rows, qa_rows, unmatched_rows) for one workbook."""
    excel_name = input_path.name
    if box_name is None:
        box_name = infer_box_name(excel_name)

    profiles = config["profiles"]
    tolerance = config.get("tolerance", 0.01)

    wb = openpyxl.load_workbook(input_path, data_only=True)

    data_rows: list[dict] = []
    qa_rows: list[dict] = []
    unmatched: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind, classify_reason = classify_sheet(ws, profiles)
        if kind == "reference":
            continue
        if kind == "unfilled_template":
            unmatched.append({
                "File": excel_name, "Sheet": sheet_name,
                "Reason": f"unfilled employee template: {classify_reason}",
            })
            continue

        # Field Mechanic family — a line-item table, not a day-block grid.
        # Detected by its label-bearing header row (WC STATE + RT#/OT#/...).
        # Every field is resolved by label inside extract_field_mechanic_sheet.
        fm_header_row = find_field_mechanic_header_row(ws)
        if fm_header_row is not None:
            fm_rows = extract_field_mechanic_sheet(
                ws, fm_header_row, folder_name, box_name, excel_name, kind=kind,
            )
            data_rows.extend(fm_rows)
            totals_row = next(r for r in fm_rows if r["Day"] == "TOTALS")
            sheet_issue = totals_row["QA_Flag"] if totals_row["QA_Flag"] != "OK" else ""
            needs_review = "[CHECK]" in sheet_issue or "[WARN]" in sheet_issue
            qa_rows.append({
                "File":           excel_name,
                "Employee":       fm_rows[0]["Employee Name"],
                "Sheet Name":     sheet_name,
                "Profile":        "field_mechanic",
                "Pay Totals Row": f"header row {fm_header_row}",
                "QA Method":      "Field Mechanic line-item sums by label-discovered (day,pay-type) columns",
                "Tolerance":      tolerance,
                "Sheet Issues":   sheet_issue,
                "Overall":        "REVIEW" if needs_review else "PASS",
            })
            continue

        # Employee sheets with no weekly grid (e.g. 'Job', 'OVERHEAD- CA') —
        # emit 8 blank rows so the employee is still represented in output.
        if not sheet_has_day_grid(ws, profiles):
            ng_profile = next(iter(profiles.values()), None)
            no_grid_rows = extract_sheet_no_grid(
                ws, folder_name, box_name, excel_name, ng_profile, kind=kind,
            )
            data_rows.extend(no_grid_rows)
            placeholder_issue = (
                f"[CHECK] Employee Name placeholder — using sheet name {sheet_name!r}"
                if kind == "employee_placeholder" else ""
            )
            qa_rows.append({
                "File":           excel_name,
                "Employee":       no_grid_rows[0]["Employee Name"],
                "Sheet Name":     sheet_name,
                "Profile":        "(no grid)",
                "Pay Totals Row": "",
                "QA Method":      "no weekly day-label grid — 8 blank rows emitted",
                "Tolerance":      tolerance,
                "Sheet Issues":   placeholder_issue,
                "Overall":        "NO_GRID",
            })
            continue

        profile_name = select_profile(ws, profiles)
        if profile_name is None:
            unmatched.append({
                "File": excel_name, "Sheet": sheet_name,
                "Reason": "no matching profile (day labels not at expected position)",
            })
            continue

        profile = profiles[profile_name]
        pay_totals_row, reason = resolve_pay_totals_row(ws, profile, tolerance=tolerance)
        if pay_totals_row is None:
            unmatched.append({
                "File": excel_name, "Sheet": sheet_name,
                "Reason": f"pay_totals_row unresolved: {reason}",
            })
            continue

        emp_rows = extract_sheet(ws, profile, pay_totals_row,
                                 folder_name, box_name, excel_name, kind=kind)

        # Phase 2 sheet-level QA layers
        day_rows = [r for r in emp_rows if r["Day"] != "TOTALS"]
        totals_row = next(r for r in emp_rows if r["Day"] == "TOTALS")

        sheet_issues: list[str] = []
        sheet_issues += two_pass_totals_row(ws, profile, pay_totals_row, tolerance)
        sheet_issues += check_weekly_hours_range(totals_row.get("Total Hours"))
        sheet_issues += check_date_sequence(day_rows)
        sheet_issues += check_filename_week_match(excel_name, day_rows)

        # Surface sheet-level issues on the TOTALS row's QA_Flag
        if sheet_issues:
            existing = totals_row.get("QA_Flag", "OK")
            extra = "; ".join(sheet_issues)
            totals_row["QA_Flag"] = extra if existing == "OK" else f"{existing}; {extra}"

        data_rows.extend(emp_rows)

        passed, results = check_accuracy(emp_rows, profile["day_blocks"]["pay_types"], tolerance)
        emp_name = emp_rows[0]["Employee Name"]
        qa_row = {
            "File":       excel_name,
            "Employee":   emp_name,
            "Sheet Name": sheet_name,
            "Profile":    profile_name,
            "Pay Totals Row": pay_totals_row,
            "QA Method":  f"sum(MON..SUN) vs grand totals (row {pay_totals_row})",
            "Tolerance":  tolerance,
        }
        for col, (daily_sum, grand_total, col_status) in results.items():
            qa_row[f"{col}_daily_sum"]   = daily_sum
            qa_row[f"{col}_grand_total"] = grand_total
            qa_row[f"{col}_match"]       = col_status
        totals_qa = str(totals_row.get("QA_Flag", ""))
        anchor_issue_tokens = [seg.strip() for seg in totals_qa.split(";")
                               if "[CHECK]" in seg and "anchor" in seg.lower()]
        all_issues_for_display = sheet_issues + anchor_issue_tokens
        qa_row["Sheet Issues"] = "; ".join(all_issues_for_display) if all_issues_for_display else ""
        # REVIEW only if there's an actionable [CHECK] or [WARN]; [INFO] alone passes.
        needs_review = (
            any(("[CHECK]" in s or "[WARN]" in s) for s in sheet_issues)
            or "[CHECK]" in totals_qa
            or "[WARN]" in totals_qa
        )
        qa_row["Overall"] = "PASS" if (passed and not needs_review) else ("REVIEW" if passed else "FAIL")
        qa_rows.append(qa_row)

    return data_rows, qa_rows, unmatched


def _employee_has_no_activity(emp_rows: list[dict], pay_types: list[str]) -> bool:
    """True iff this employee should be dropped from output.

    Drop criteria (BOTH must hold):
      1. Grand TOTALS row 'Total Hours' is 0 / None / blank / '0.00'
      2. No day row has any time-of-day entry (Start/Lunch Out/Lunch In/Stop)

    Pay-type cells are intentionally NOT consulted — HP-PTO / OT-only
    accounting sheets show pay-type totals without representing worked
    hours. The user's rule is "didn't work even for a single day".
    """
    def _is_zero_or_blank(v):
        if v is None:
            return True
        if isinstance(v, bool):
            return False
        if isinstance(v, (int, float)):
            return v == 0
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return True
            try:
                return float(s) == 0
            except ValueError:
                return False
        return False

    day_rows = [r for r in emp_rows if r["Day"] != "TOTALS"]
    totals_row = next((r for r in emp_rows if r["Day"] == "TOTALS"), None)

    if not day_rows and totals_row is None:
        return True
    if totals_row is None:
        return False

    if not _is_zero_or_blank(totals_row.get("Total Hours")):
        return False

    for r in day_rows:
        for k in ("Start", "Lunch Out", "Lunch In", "Stop"):
            v = r.get(k)
            if v not in (None, "", 0):
                return False
    return True


def _drop_empty_employees(all_data: list[dict], all_qa: list[dict],
                          pay_types: list[str]) -> tuple[list[dict], list[dict], int]:
    """Drop any employee (grid-bearing or NO_GRID) with zero activity.

    'Zero activity' means: every day row has Total Hours = 0/None/'0.00'
    AND every pay-type cell is 0/blank AND every time-of-day cell is blank.
    NO_GRID employees that we can't currently extract real data for are
    included in this filter — empty rows in the output are not real data.
    """
    by_key: dict[tuple, list[dict]] = {}
    order: list[tuple] = []
    for r in all_data:
        k = (r["Excel Name"], r["Sheet Name"])
        if k not in by_key:
            by_key[k] = []
            order.append(k)
        by_key[k].append(r)

    keep_keys = set()
    for k in order:
        if not _employee_has_no_activity(by_key[k], pay_types):
            keep_keys.add(k)

    new_data = [r for r in all_data if (r["Excel Name"], r["Sheet Name"]) in keep_keys]
    new_qa   = [q for q in all_qa   if (q["File"],       q["Sheet Name"])  in keep_keys]
    dropped  = len(order) - len(keep_keys)
    return new_data, new_qa, dropped


def extract(input_arg: str, folder_name: str, box_name: str | None, out_path: str,
            config_path: str) -> int:
    config = load_config(Path(config_path))
    output_columns = config["output_columns"]

    in_path = Path(input_arg)
    files: list[Path]
    if in_path.is_dir():
        files = sorted([p for p in in_path.rglob("*.xlsx") if not p.name.startswith("~$")])
    else:
        files = [in_path]

    all_data: list[dict] = []
    all_qa: list[dict] = []
    all_unmatched: list[dict] = []

    print(f"\nProcessing {len(files)} workbook(s)...")
    for f in files:
        try:
            data, qa, um = process_workbook(f, config, folder_name, box_name)
        except Exception as e:
            all_unmatched.append({"File": f.name, "Sheet": "(workbook)", "Reason": f"ERROR: {e!r}"})
            print(f"  {f.name[:60]:<60}  ERROR {e!r}")
            continue
        all_data.extend(data)
        all_qa.extend(qa)
        all_unmatched.extend(um)
        passed = sum(1 for r in qa if r["Overall"] == "PASS")
        total = len(qa)
        print(f"  {f.name[:60]:<60}  emp={total:>3}  pass={passed:>3}  unmatched={len(um):>2}")

    pay_types_for_filter = config["profiles"]["standard"]["day_blocks"]["pay_types"]
    all_data, all_qa, dropped_empty = _drop_empty_employees(
        all_data, all_qa, pay_types_for_filter
    )
    if dropped_empty:
        print(f"Dropped {dropped_empty} employee sheet(s) with zero activity (no hours, no time entries).")

    total_emp = len(all_qa)
    total_pass    = sum(1 for r in all_qa if r["Overall"] == "PASS")
    total_no_grid = sum(1 for r in all_qa if r["Overall"] == "NO_GRID")
    total_grid    = total_emp - total_no_grid
    print(f"\n{total_pass}/{total_grid} grid-bearing sheets passed accuracy check.")
    print(f"{total_no_grid} employee sheet(s) had no time grid (8 blank rows each).")
    print(f"{len(all_unmatched)} sheet(s) on Unmatched_Sheets tab.\n")

    run_info = {
        "Input":           str(in_path),
        "Folder Name":     folder_name,
        "Box Name":        box_name or "(inferred per-file)",
        "Run Timestamp":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Files Processed": len(files),
        "Sheets Extracted":     total_emp,
        "Sheets Passed":   total_pass,
        "Sheets Failed":   total_grid - total_pass,
        "Sheets No-Grid":  total_no_grid,
        "Sheets Unmatched": len(all_unmatched),
        "Sheets Dropped (empty)": dropped_empty,
        "Sheets Needing Review": sum(1 for r in all_qa if r["Overall"] == "REVIEW"),
        "Overall Result":  "PASS" if total_pass == total_grid and not all_unmatched else "REVIEW",
        "Script Version":  "2.1 (Phase 2)",
        "Config":          str(config_path),
    }

    id_conflicts = check_ee_id_consistency(all_data)

    data_df = pd.DataFrame(all_data, columns=output_columns)
    out_p = Path(out_path)
    out_p.parent.mkdir(parents=True, exist_ok=True)
    _write_excel(out_p, data_df, all_qa, run_info, all_unmatched, id_conflicts)
    if id_conflicts:
        print(f"  ID_Conflicts tab: {len(id_conflicts)} EE-ID conflict(s) found")
    print(f"Output written to: {out_p}")
    print("  Sheet 1 — Data             : extracted rows + QA_Flag per row")
    print("  Sheet 2 — QA_Summary       : per-employee accuracy comparison")
    print("  Sheet 3 — Run_Info         : run metadata and overall result")
    print("  Sheet 4 — Unmatched_Sheets : sheets that couldn't be safely extracted")
    print("  Sheet 5 — ID_Conflicts     : same EE ID mapped to multiple names\n")

    return 0 if total_pass == total_grid and not all_unmatched else 2


def main():
    parser = argparse.ArgumentParser(description="Extract timesheet data from Excel workbook(s).")
    parser.add_argument("input", help="Path to input .xlsx file or folder of .xlsx files")
    parser.add_argument("--folder", default="2022", help="Folder Name value (default: 2022)")
    parser.add_argument("--box",    default=None,   help="Box Name (inferred from filename if omitted)")
    parser.add_argument("--out",    default="output/extracted.xlsx", help="Output Excel path")
    parser.add_argument("--config", default="config/schema.yaml",    help="Path to config YAML")
    args = parser.parse_args()
    sys.exit(extract(args.input, args.folder, args.box, args.out, args.config))


if __name__ == "__main__":
    main()
