"""Inspect the user's feedback file to see what was annotated."""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
FB = ROOT / "phase2_corpus_v3 - Feedback.xlsx"

wb = load_workbook(FB, data_only=True)
print(f"Sheets in feedback file: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n--- Sheet '{sn}' — {ws.max_row} rows x {ws.max_column} cols ---")
wb.close()

# Read the Data sheet and look for any annotated rows (likely a new column)
for sn in ["Data", "QA_Summary"]:
    df = pd.read_excel(FB, sheet_name=sn)
    print(f"\n=== Sheet '{sn}' columns ===")
    print(list(df.columns))
    print(f"Rows: {len(df)}")
