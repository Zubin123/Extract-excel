"""Verify the 6 user claims against fresh extraction output."""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
OUT = ROOT / "output" / "verify_run.xlsx"
DATA = ROOT / "data"

data_df = pd.read_excel(OUT, sheet_name="Data")
qa_df = pd.read_excel(OUT, sheet_name="QA_Summary")
unmatched_df = pd.read_excel(OUT, sheet_name="Unmatched_Sheets")


def find_file(name_substr):
    matches = list(DATA.rglob(f"*{name_substr}*.xlsx"))
    return matches[0] if matches else None


def input_sheets(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    out = []
    for ws in wb.worksheets:
        a1 = ws["A1"].value
        b2 = ws["B2"].value
        out.append({"sheet": ws.title, "A1": a1, "B2": b2})
    wb.close()
    return out


def output_sheets_for(excel_name):
    rows = data_df[data_df["Excel Name"] == excel_name]
    return sorted(rows["Sheet Name"].dropna().unique().tolist())


def banner(title):
    print("\n" + "=" * 88)
    print(title)
    print("=" * 88)


# ---------- CLAIM 1 ----------
banner("CLAIM 1: WE 12 18 21 CA 6427 - UPLOADED - missing Delgado/Griego/Nunez/McDonald/Benedict/Rivera")
fname = "WE 12 18 21 CA 6427"
xlsx = find_file(fname)
print(f"Input: {xlsx.name}")
input_list = input_sheets(xlsx)
print(f"Total sheets in workbook: {len(input_list)}")
sheets_in_output = output_sheets_for(xlsx.name)
print(f"Extracted sheets ({len(sheets_in_output)}): {sheets_in_output}")
print("\nAll input sheets and their A1/B2 anchors:")
for s in input_list:
    extracted = "EXTRACTED" if s["sheet"] in sheets_in_output else "*** MISSED ***"
    print(f"  [{extracted:14}] {s['sheet']!r:30}  A1={s['A1']!r:30}  B2={s['B2']!r}")
for claimed in ["Delgado", "Griego", "Nunez", "McDonald", "Benedict", "Rivera"]:
    found = [s for s in input_list if claimed.lower() in str(s["sheet"]).lower()
             or claimed.lower() in str(s["A1"]).lower()]
    if found:
        for f in found:
            in_out = f["sheet"] in sheets_in_output
            print(f"  -> Claim '{claimed}': sheet {f['sheet']!r} extracted={in_out}")
    else:
        print(f"  -> Claim '{claimed}': NO SHEET FOUND MATCHING THIS NAME")


# ---------- CLAIM 2 ----------
banner("CLAIM 2: WE 12 18 21 Hall - OVERHEAD-CA missed")
fname = "WE 12 18 21 Hall"
xlsx = find_file(fname)
print(f"Input: {xlsx.name}")
input_list = input_sheets(xlsx)
sheets_in_output = output_sheets_for(xlsx.name)
print(f"Total sheets: {len(input_list)}  /  Extracted: {len(sheets_in_output)}")
for s in input_list:
    extracted = "EXTRACTED" if s["sheet"] in sheets_in_output else "*** MISSED ***"
    print(f"  [{extracted:14}] {s['sheet']!r:35}  A1={s['A1']!r:30}  B2={s['B2']!r}")


# ---------- CLAIM 3 ----------
banner("CLAIM 3: WE 12 25 21 Hall - OVERHEAD-CA missed")
fname = "WE 12 25 21 Hall"
xlsx = find_file(fname)
print(f"Input: {xlsx.name}")
input_list = input_sheets(xlsx)
sheets_in_output = output_sheets_for(xlsx.name)
print(f"Total sheets: {len(input_list)}  /  Extracted: {len(sheets_in_output)}")
for s in input_list:
    extracted = "EXTRACTED" if s["sheet"] in sheets_in_output else "*** MISSED ***"
    print(f"  [{extracted:14}] {s['sheet']!r:35}  A1={s['A1']!r:30}  B2={s['B2']!r}")


# ---------- CLAIM 4 ----------
banner("CLAIM 4: WE 12 25 21 6562 UPL -WA 4937-5257-4377 - Hanson missed")
fname = "WE 12 25 21 6562 UPL -WA 4937"
xlsx = find_file(fname)
print(f"Input: {xlsx.name}")
input_list = input_sheets(xlsx)
sheets_in_output = output_sheets_for(xlsx.name)
print(f"Total sheets: {len(input_list)}  /  Extracted: {len(sheets_in_output)}")
print("\nSheets containing 'Hanson':")
for s in input_list:
    if "hanson" in str(s["sheet"]).lower() or "hanson" in str(s["A1"]).lower():
        in_out = s["sheet"] in sheets_in_output
        print(f"  sheet={s['sheet']!r} A1={s['A1']!r} B2={s['B2']!r} extracted={in_out}")
if not any("hanson" in str(s["sheet"]).lower() or "hanson" in str(s["A1"]).lower() for s in input_list):
    print("  NO HANSON SHEET IN THIS WORKBOOK")
print("\nAll sheets in this workbook:")
for s in input_list:
    extracted = "EXTRACTED" if s["sheet"] in sheets_in_output else "*** MISSED ***"
    print(f"  [{extracted:14}] {s['sheet']!r:35}  A1={s['A1']!r:30}  B2={s['B2']!r}")


# ---------- CLAIM 5 ----------
banner("CLAIM 5: WE 12 25 21 6818 UPL - WC State extracted from wrong column (FSL instead of WC State)")
fname = "WE 12 25 21 6818"
xlsx = find_file(fname)
print(f"Input: {xlsx.name}")
wb = load_workbook(xlsx, data_only=True, read_only=False)
# Inspect first employee sheet's header area around row 13
for ws in wb.worksheets:
    a1 = ws["A1"].value
    b2 = ws["B2"].value
    if isinstance(a1, str) and a1.strip() and isinstance(b2, int):
        print(f"\n--- Sheet {ws.title!r} (A1={a1!r}, EE={b2}) ---")
        # Print row 12 (likely headers) and row 13 (likely values) across cols A..M
        for r in [12, 13]:
            cells = [(ws.cell(row=r, column=c).coordinate,
                      ws.cell(row=r, column=c).value) for c in range(1, 14)]
            print(f"  Row {r}: {cells}")
        break
wb.close()
# What did we extract?
rows = data_df[data_df["Excel Name"] == xlsx.name].head(8)
print(f"\nExtracted (first 8 rows) for {xlsx.name}:")
print(rows[["Sheet Name", "Employee Name", "EE ID", "WC State", "ST"]].to_string())


# ---------- CLAIM 6 ----------
banner("CLAIM 6: WE 12 18 21 & WE 12 25 21 - WC State missing across multiple files")
# Show distribution of WC State by Excel Name
wc_summary = (data_df.groupby("Excel Name")
              .agg(rows=("WC State", "size"),
                   wc_nonblank=("WC State", lambda s: s.notna().sum()),
                   wc_values=("WC State", lambda s: ",".join(sorted(s.dropna().astype(str).unique()))))
              .reset_index())
wc_summary["wc_blank_pct"] = (1 - wc_summary["wc_nonblank"] / wc_summary["rows"]) * 100
print("\nFiles where WC State is blank/missing for any row:")
problem = wc_summary[wc_summary["wc_blank_pct"] > 0].sort_values("wc_blank_pct", ascending=False)
print(problem.to_string(index=False))

print("\nUnique WC State values across all output:")
print(data_df["WC State"].value_counts(dropna=False).to_string())
