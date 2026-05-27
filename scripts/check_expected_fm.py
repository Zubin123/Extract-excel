"""For the 6 Field Mechanic NO_GRID sheets, check what the expected (right
side) block in the feedback file says — does it have data or is it blank?"""
from pathlib import Path
import openpyxl
ROOT = Path(__file__).resolve().parents[1]
fb = openpyxl.load_workbook(ROOT / "phase2_corpus_v4 - Feedback.xlsx", data_only=True)
ws = fb["Data"]

FM_SHEETS = {
    ("WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx", "Job"),
    ("WE 12 25 21 Hall 4906-2066-0647_1.xlsx", "JOB"),
    ("WE 12 25 21 Juan Medina 4920-9705-5655_1.xlsx", "Job"),
    ("WE 12 25 21 Michael Beeler 4902-6840-4647_1.xlsx", "Job"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "JOB"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "OVERHEAD - CA"),
}

# Column layout:
# 1-25 ours; 27-48 expected (Folder Name at 27, Excel at 28, Sheet at 29,
# EmpName at 30, EE ID at 31, WC at 32, ST at 33, f at 34, Day at 35,
# Start at 36, Lunch Out 37, Lunch In 38, Stop 39, TH 40, ...)
# Wait — earlier I saw "f" at col 34. Let me verify with r2.

for r in range(2, ws.max_row + 1):
    our_file = ws.cell(row=r, column=3).value
    our_sheet = ws.cell(row=r, column=4).value
    key = (our_file, our_sheet)
    if key not in FM_SHEETS:
        continue
    our_day = ws.cell(row=r, column=10).value
    # Read expected block — cols 28-40
    exp_file = ws.cell(row=r, column=28).value
    exp_sheet = ws.cell(row=r, column=29).value
    exp_emp = ws.cell(row=r, column=30).value
    exp_day = ws.cell(row=r, column=35).value
    exp_start = ws.cell(row=r, column=36).value
    exp_stop = ws.cell(row=r, column=39).value
    exp_th = ws.cell(row=r, column=40).value
    print(f"row {r}: {our_file[-30:]} :: {our_sheet:<15} our_day={our_day}")
    print(f"        expected: file={exp_file!r} sheet={exp_sheet!r} emp={exp_emp!r}")
    print(f"        expected day={exp_day} start={exp_start} stop={exp_stop} TH={exp_th}")
    print()
