"""Open phase2_corpus_v4 - Feedback.xlsx and dump its structure / annotations."""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
p = ROOT / "phase2_corpus_v4 - Feedback.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
print(f"Sheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== {sn!r} ({ws.max_row} rows x {ws.max_column} cols) ===")
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print("Headers:", headers)
    # First 5 rows
    for r in range(2, min(7, ws.max_row + 1)):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  r{r}: {row}")
