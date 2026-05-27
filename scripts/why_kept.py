"""For each of the 17 listed sheets, dump what cells the drop filter sees."""
from pathlib import Path
import openpyxl
ROOT = Path(__file__).resolve().parents[1]
wb = openpyxl.load_workbook(ROOT / "output" / "phase2_corpus_v4.xlsx", data_only=True)
ws = wb["Data"]
hdrs = [c.value for c in ws[1]]

TARGETS = {
    ("WE 12 25 21 6562 UPL -WA 4937-5257-4377_1.xlsx", "Wilson G."),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Case"),
    ("WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx", "Job"),
}

rows_by_key = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = dict(zip(hdrs, row))
    k = (rec["Excel Name"], rec["Sheet Name"])
    if k in TARGETS:
        rows_by_key.setdefault(k, []).append(rec)

for k, rows in rows_by_key.items():
    print(f"=== {k} ===")
    for r in rows:
        relevant = {fld: r.get(fld) for fld in (
            "Day", "Date", "Start", "Lunch Out", "Lunch In", "Stop",
            "Total Hours", "RT", "OT", "DT", "PD", "4D", "4A"
        )}
        print(f"  {relevant}")
    print()
