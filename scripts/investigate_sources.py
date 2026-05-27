"""Find where the expected employee names and WC State values actually live in the source."""
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
DATA = ROOT / "data"


def show_sheet_top(xlsx, sheet_name, rows=15, cols=12):
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet_name]
    print(f"\n--- {xlsx.name} / sheet '{sheet_name}' (first {rows} rows × {cols} cols) ---")
    for r in range(1, rows + 1):
        line = []
        for c in range(1, cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            col_letter = chr(ord("A") + c - 1)
            line.append(f"{col_letter}{r}={v!r}")
        if line:
            print(f"  row {r}: " + " | ".join(line))
    wb.close()


# Hall files — find where Peter Joe Hall's name might live
hall_files = list(DATA.rglob("*Hall*.xlsx"))
print("HALL FILES:")
for f in hall_files:
    print(f"\n{'='*88}\n{f.name}\n{'='*88}")
    wb = load_workbook(f, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    wb.close()
    # Look at the Employees sheet
    show_sheet_top(f, "Employees", rows=10, cols=10)
    show_sheet_top(f, "OVERHEAD - CA", rows=15, cols=12)


# Ibarra — why is WC State blank?
print("\n\n" + "="*88)
print("IBARRA SHEET — full row 12 and 13")
print("="*88)
ibarra_file = next(DATA.rglob("*CA 6427 - UPLOADED*4901*.xlsx"))
wb = load_workbook(ibarra_file, data_only=True)
ws = wb["Ibarra"]
for r in [12, 13, 14, 15]:
    line = []
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        col_letter = chr(ord("A") + c - 1) if c <= 26 else chr(ord("A") + (c-1)//26 - 1) + chr(ord("A") + (c-1) % 26)
        if v is not None:
            line.append(f"{col_letter}{r}={v!r}")
    print(f"row {r}: {' | '.join(line)}")
wb.close()

# Schnider OH — why blank WC State?
print("\n\n" + "="*88)
print("SCHNIDER OH SHEET — row 12 and 13")
print("="*88)
wb = load_workbook(ibarra_file, data_only=True)
ws = wb["Schnider OH"]
for r in [12, 13, 14, 15]:
    line = []
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        col_letter = chr(ord("A") + c - 1) if c <= 26 else chr(ord("A") + (c-1)//26 - 1) + chr(ord("A") + (c-1) % 26)
        if v is not None:
            line.append(f"{col_letter}{r}={v!r}")
    print(f"row {r}: {' | '.join(line)}")
wb.close()

# Delgado — does B2 give EE ID 3398, and does any cell in the sheet contain the real name?
print("\n\n" + "="*88)
print("DELGADO SHEET — A1, B2, and Employees lookup")
print("="*88)
wb = load_workbook(ibarra_file, data_only=True)
ws = wb["Delgado"]
print(f"A1 = {ws['A1'].value!r}")
print(f"B2 = {ws['B2'].value!r}")
# Check if there's an Employees sheet
if "Employees" in wb.sheetnames:
    emp_ws = wb["Employees"]
    print(f"\nEmployees sheet — first 30 rows:")
    for r in range(1, min(emp_ws.max_row + 1, 40)):
        line = []
        for c in range(1, 6):
            v = emp_ws.cell(row=r, column=c).value
            if v is not None:
                line.append(f"{chr(ord('A') + c - 1)}{r}={v!r}")
        if line:
            print(f"  row {r}: " + " | ".join(line))
wb.close()
