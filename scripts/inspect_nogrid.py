"""Dump raw structure of every NO_GRID-classified employee sheet.

For each, we print:
  - identity cells (A1, B1, B2, C2)
  - any cell in rows 1-15 that contains a day-name token (MON/MONDAY/TUE/...)
  - any cell in rows 1-15 that contains 'Start' / 'Stop' / 'Lunch' (time labels)
  - the row that has the most date-typed cells (likely the date row)

This bypasses our classifier entirely so we can eyeball whether the sheet
really has no weekly grid or whether the grid is in an unexpected location.
"""
from __future__ import annotations
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from datetime import datetime, date
from anchors import classify_sheet, sheet_has_day_grid, col_index_to_letter

import yaml
cfg = yaml.safe_load((ROOT / "config" / "schema.yaml").read_text())
profiles = cfg["profiles"]

DAY_FULL = {"MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY"}
DAY_SHORT = {"MON","TUE","WED","THUR","THU","FRI","SAT","SUN"}
TIME_LABELS = {"START","STOP","LUNCH OUT","LUNCH IN"}

def scan_sheet(ws):
    day_hits = []
    time_hits = []
    date_row_counts = {}
    max_r = min(ws.max_row or 0, 20)
    max_c = min(ws.max_column or 0, 80)
    for r in range(1, max_r + 1):
        date_count = 0
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u in DAY_FULL or u in DAY_SHORT:
                    day_hits.append((r, c, v))
                if u in TIME_LABELS:
                    time_hits.append((r, c, v))
            if isinstance(v, (datetime, date)):
                date_count += 1
        if date_count >= 3:
            date_row_counts[r] = date_count
    return day_hits, time_hits, date_row_counts


def main():
    files = sorted((ROOT / "data").rglob("*.xlsx"))
    files = [p for p in files if not p.name.startswith("~$") and "Header Template" not in p.name]

    nogrid_sheets = []
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            kind, _ = classify_sheet(ws, profiles)
            if kind in ("reference", "unfilled_template"):
                continue
            if kind in ("employee", "employee_placeholder", "employee_alt_layout"):
                if not sheet_has_day_grid(ws, profiles):
                    nogrid_sheets.append((f, sn, ws, kind))

    print(f"Found {len(nogrid_sheets)} NO_GRID employee sheets.\n")

    for (f, sn, ws, kind) in nogrid_sheets:
        a1 = ws["A1"].value
        b1 = ws["B1"].value
        b2 = ws["B2"].value
        c2 = ws["C2"].value
        day_hits, time_hits, date_rows = scan_sheet(ws)
        print(f"--- {f.name} :: {sn!r}  ({kind})")
        print(f"    A1={a1!r}  B1={b1!r}  B2={b2!r}  C2={c2!r}")
        if day_hits:
            print(f"    DAY tokens found: {[(col_index_to_letter(c)+str(r), v) for (r,c,v) in day_hits]}")
        else:
            print(f"    DAY tokens: NONE")
        if time_hits:
            print(f"    TIME labels found: {[(col_index_to_letter(c)+str(r), v) for (r,c,v) in time_hits]}")
        else:
            print(f"    TIME labels: NONE")
        if date_rows:
            print(f"    Date-heavy rows: {date_rows}")
        print()


if __name__ == "__main__":
    main()
