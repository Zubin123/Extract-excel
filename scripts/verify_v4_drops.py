"""Verify the 17 listed empty employees are gone from v4, and check that
no real-work-having employee was dropped by mistake.
"""
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
wb = openpyxl.load_workbook(ROOT / "output" / "phase2_corpus_v4.xlsx", data_only=True)
ws = wb["Data"]
hdrs = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(hdrs)}

EMPTY_EMP_SHEETS = {
    ("WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx", "Job"),
    ("WE 12 25 21 Hall 4906-2066-0647_1.xlsx", "JOB"),
    ("WE 12 25 21 Juan Medina 4920-9705-5655_1.xlsx", "Job"),
    ("WE 12 25 21 Michael Beeler 4902-6840-4647_1.xlsx", "Job"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "JOB"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "OVERHEAD - CA"),
    ("WE 12 25 21 6562 UPL -WA 4937-5257-4377_1.xlsx", "Wilson G."),
    ("WE 12 25 21 6562 UPL -WA 4937-5257-4377_1.xlsx", "Jackson HP-PTO"),
    ("WE 12 25 21 6799 McIntire UPL -WA 4919-9109-7769_1.xlsx", "McIntire"),
    ("WE 12 25 21 6818 UPL -WA 4903-4699-6137_1.xlsx", "I Christopherson PTO-HP"),
    ("WE 12 25 21 CA 6721 UPL -WA 4921-9281-7577_1.xlsx", "David Cortez PTO-HP"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Case"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Sanders"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Eric Schmidt"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Raper  HP-PTO"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Peake "),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Zeulner"),
}

# Collect all (file, sheet) keys still present in v4
present = set()
sheet_row_counts = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = dict(zip(hdrs, row))
    key = (rec["Excel Name"], rec["Sheet Name"])
    present.add(key)
    sheet_row_counts[key] = sheet_row_counts.get(key, 0) + 1

print(f"Total employee sheets in v4 Data: {len(present)}")
print(f"Total rows in v4 Data: {sum(sheet_row_counts.values())}")
print()
print("=== Check: listed 17 empty sheets should be GONE ===")
still_present = EMPTY_EMP_SHEETS & present
missing = EMPTY_EMP_SHEETS - present
print(f"  Still present (should be 0): {len(still_present)}")
for k in still_present:
    print(f"    !! STILL THERE: {k}")
print(f"  Successfully dropped: {len(missing)}/{len(EMPTY_EMP_SHEETS)}")

# Also: scan QA_Summary for any kept sheet with Total Hours = 0/blank to spot-check
ws2 = wb["QA_Summary"]
hdrs2 = [c.value for c in ws2[1]]
zero_total_kept = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    rec = dict(zip(hdrs2, row))
    th = rec.get("Total Hours_grand_total")
    if th in (0, 0.0, "0", "0.00", None):
        zero_total_kept.append((rec["File"], rec["Sheet Name"], th, rec["Overall"]))
print()
print(f"=== Sheets still in v4 with grand Total Hours = 0/None: {len(zero_total_kept)} ===")
for f, s, th, ov in zero_total_kept[:20]:
    print(f"  {f[-40:]:<40} :: {s:<20} TH={th!r} Overall={ov}")
