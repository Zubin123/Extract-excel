"""One-off probe: find where Total Hours lives and look for a grand-total cell."""
import openpyxl
from pathlib import Path

INPUT = Path(__file__).parent.parent / "data" / "WE 01 08 22 CA GF'S - WA 4939-0899-6263_1.xlsx"
wb = openpyxl.load_workbook(INPUT, data_only=True)

sheets_to_check = ["Case", "Sanders", "Schmidt", "Raper ", "Cutler", "Covey"]
day_cols = ["L", "R", "X", "AD", "AJ", "AP", "AV"]
days     = ["MON", "TUE", "WED", "THUR", "FRI", "SAT", "SUN"]

for sn in sheets_to_check:
    if sn not in wb.sheetnames:
        continue
    ws = wb[sn]
    print(f"--- {sn!r} ---")
    for d, c in zip(days, day_cols):
        print(f"  {d} row9 ({c}9): {ws[c + '9'].value!r}")
    # Look right of BG for a possible grand-total cell on row 9 and row 24
    for col_letter in ("BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ"):
        print(f"  {col_letter}9: {ws[col_letter + '9'].value!r}   {col_letter}24: {ws[col_letter + '24'].value!r}")
    print()
