"""Check whether the Employees roster sheet has the EE IDs we need to look up."""
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
DATA = ROOT / "data"


def build_roster(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    if "Employees" not in wb.sheetnames:
        print(f"  {xlsx.name}: no Employees sheet")
        wb.close()
        return {}
    ws = wb["Employees"]
    # Find header row (look for 'EE ID' label in first column or two)
    header_row = None
    for r in range(1, 10):
        for c in range(1, 5):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() == "EE ID":
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        print(f"  {xlsx.name}: Employees sheet has no 'EE ID' header")
        wb.close()
        return {}

    # Find column positions for ID, Name, and any state columns
    headers = {}
    for c in range(1, 12):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            headers[v.strip().upper()] = c
    id_col = headers.get("EE ID")
    name_col = headers.get("EMPLOYEE NAME")
    wc_col = headers.get("WORKERS' COMP STATE") or headers.get("WC STATE")
    st_col = headers.get("WORK STATE") or headers.get("ST")

    roster = {}
    for r in range(header_row + 1, ws.max_row + 1):
        ee = ws.cell(row=r, column=id_col).value if id_col else None
        if not isinstance(ee, int):
            continue
        name = ws.cell(row=r, column=name_col).value if name_col else None
        wc = ws.cell(row=r, column=wc_col).value if wc_col else None
        st = ws.cell(row=r, column=st_col).value if st_col else None
        roster[ee] = {"name": name, "wc_state": wc, "st": st}
    wb.close()
    return roster


# Check the 7 sheets that need lookup
targets = {
    "WE 12 18 21 CA 6427": [3398, 4769, 8170, 7306, 1623, 8917],  # Delgado, Griego, Nunez, McDonald, Benedict, Rivera
    "WE 12 25 21 6562 UPL -WA 4937": [4961],                       # Hanson
    "WE 12 18 21 Hall 4926": [4876],                               # Peter Joe Hall
    "WE 12 25 21 Hall 4906": [4876],
}

for tag, ids in targets.items():
    print(f"\n=== {tag} ===")
    matches = list(DATA.rglob(f"*{tag}*.xlsx"))
    if not matches:
        print("  file not found")
        continue
    xlsx = matches[0]
    roster = build_roster(xlsx)
    print(f"  Roster size: {len(roster)} entries")
    for ee in ids:
        entry = roster.get(ee)
        print(f"  EE {ee}: {entry}")
