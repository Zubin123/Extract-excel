"""Find every #VALUE! in row 9 and show what's in rows 5-8 for that day."""
import openpyxl
from pathlib import Path

INPUT = Path(__file__).parent.parent / "data" / "WE 01 08 22 CA GF'S - WA 4939-0899-6263_1.xlsx"

REFERENCE_SHEETS = {
    "Master Template (9) JOB", "Certified Codes", "Employees",
    "Job Details", "Job List", "ColumnLists",
}

DAY_BLOCKS = [
    ("MON",  "L"),  ("TUE",  "R"),  ("WED",  "X"),  ("THUR", "AD"),
    ("FRI",  "AJ"), ("SAT",  "AP"), ("SUN",  "AV"),
]

# Load both with computed values, and with formulas (data_only=False) to compare.
wb_data = openpyxl.load_workbook(INPUT, data_only=True)
wb_form = openpyxl.load_workbook(INPUT, data_only=False)

print(f"{'Sheet':<14} {'Day':<5} {'Start':<14} {'LunchOut':<14} {'LunchIn':<14} {'Stop':<14} {'TotalHrs':<10}")
print("-" * 90)
for sn in wb_data.sheetnames:
    if sn in REFERENCE_SHEETS:
        continue
    ws_d = wb_data[sn]
    ws_f = wb_form[sn]
    for day, col in DAY_BLOCKS:
        total = ws_d[col + "9"].value
        start = ws_d[col + "5"].value
        lout  = ws_d[col + "6"].value
        lin   = ws_d[col + "7"].value
        stop  = ws_d[col + "8"].value
        is_value_err = isinstance(total, str) and total.startswith("#")
        is_blank_or_text = not isinstance(start, (int, float)) and start is not None
        if is_value_err or is_blank_or_text:
            print(f"{sn!r:<14} {day:<5} {str(start)[:13]:<14} {str(lout)[:13]:<14} {str(lin)[:13]:<14} {str(stop)[:13]:<14} {str(total)[:9]:<10}")

print("\n=== BB9 grand totals per sheet ===")
for sn in wb_data.sheetnames:
    if sn in REFERENCE_SHEETS:
        continue
    ws_d = wb_data[sn]
    ws_f = wb_form[sn]
    bb9_val = ws_d["BB9"].value
    bb9_formula = ws_f["BB9"].value
    print(f"  {sn!r:<14}: BB9 cached = {bb9_val!r:<15}  formula = {bb9_formula!r}")
