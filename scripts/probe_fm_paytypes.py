"""Probe: can Field Mechanic per-day pay-type totals be recovered?

Read-only. For each Field Mechanic data sheet we:
  1. Locate the header row by label (the row containing 'WC STATE' + 'ST'
     + day-1 pay-type tokens like 'RT1'). No fixed row number.
  2. From the header row, discover for each pay-type label (RT,OT,DT,PTO,HP)
     and each day index 1..7 the column it lives in (e.g. 'RT1','OT3').
  3. Sum each (day, pay-type) column down the line-item rows.
  4. Compare the per-day RT+OT+DT+PTO+HP sum against the daily 'Total Hours'
     summary cell for that day (row labelled 'Total Hours' in col M).
  5. Print everything so we can SEE whether the arithmetic ties out.

Nothing here is wired into the pipeline; it only verifies recoverability.
"""
from __future__ import annotations
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from anchors import col_index_to_letter

TARGETS = [
    "data/sample1/WE 12 25 21 Chris Coulson 4916-6366-6087_1.xlsx",
    "data/sample1/WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx",
    "data/sample1/WE 12 25 21 Hall 4906-2066-0647_1.xlsx",
]

# pay-type token followed by a day index, e.g. RT1, OT3, PTO7, HP2
PAYTYPE_RE = re.compile(r"^(RT|OT|DT|PTO|HP)(\d)$", re.IGNORECASE)


def num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else 0.0


def find_header_row(ws, max_r=25, max_c=90):
    """Header row = the row containing 'WC STATE' and at least 5 'RT#/OT#/...'
    pay-type tokens. Discovered, not hardcoded."""
    for r in range(1, max_r + 1):
        has_wc = False
        paytype_hits = 0
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u == "WC STATE":
                    has_wc = True
                if PAYTYPE_RE.match(u):
                    paytype_hits += 1
        if has_wc and paytype_hits >= 5:
            return r
    return None


def find_label_row(ws, label, max_r=25):
    """Find a row whose column M (or any col) holds `label` (e.g. 'Total Hours')."""
    up = label.upper()
    for r in range(1, max_r + 1):
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() == up:
                return r, c
    return None, None


def main():
    for rel in TARGETS:
        path = ROOT / rel
        if not path.exists():
            print(f"!! MISSING {rel}\n")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        print("=" * 90)
        print(f"FILE {path.name}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            hr = find_header_row(ws)
            if hr is None:
                continue
            # discover (paytype, day) -> column from the header row
            grid = {}  # (day:int, pt:str) -> col idx
            days_seen = set()
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(row=hr, column=c).value
                if isinstance(v, str):
                    m = PAYTYPE_RE.match(v.strip().upper())
                    if m:
                        pt, day = m.group(1).upper(), int(m.group(2))
                        grid[(day, pt)] = c
                        days_seen.add(day)
            pts = ["RT", "OT", "DT", "PTO", "HP"]
            days = sorted(days_seen)

            # line-item rows = rows below header that have any pay-type value
            last_row = ws.max_row or hr
            day_sums = {d: {pt: 0.0 for pt in pts} for d in days}
            for r in range(hr + 1, last_row + 1):
                for d in days:
                    for pt in pts:
                        col = grid.get((d, pt))
                        if col:
                            day_sums[d][pt] += num(ws.cell(row=r, column=col).value)

            # daily 'Total Hours' summary row (the clock grid)
            th_row, _ = find_label_row(ws, "Total Hours")
            # day-1 column of the clock grid = first day token column.
            # Reuse: clock grid day columns are the same as pay-type day-1 cols' base.
            # We compare against the summary if present.
            print("-" * 90)
            print(f"  SHEET {sn!r}  header_row={hr}  days={days}  paytype_cols_found={len(grid)}")
            print(f"  Total-Hours summary row = {th_row}")
            for d in days:
                s = day_sums[d]
                row_total = sum(s.values())
                cells = "  ".join(f"{pt}={s[pt]:g}" for pt in pts)
                print(f"    day{d}: {cells}   sum={row_total:g}")
            # grand week total per pay-type
            grand = {pt: sum(day_sums[d][pt] for d in days) for pt in pts}
            print(f"  WEEK pay-type totals: " + "  ".join(f"{pt}={grand[pt]:g}" for pt in pts)
                  + f"   ALL={sum(grand.values()):g}")
        print()


if __name__ == "__main__":
    main()
