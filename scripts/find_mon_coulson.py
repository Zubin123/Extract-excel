import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from anchors import col_index_to_letter, DAY_TOKENS

path = ROOT / "data" / "sample1" / "WE 12 25 21 Chris Coulson 4916-6366-6087_1.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for sn in ("Job", "OVERHEAD- CA"):
    ws = wb[sn]
    print(f"=== {sn} ===")
    # Scan first 30 rows for any day-token text
    for r in range(1, 31):
        hits = []
        for c in range(1, 80):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() in DAY_TOKENS:
                hits.append(f"{col_index_to_letter(c)}{r}={v!r}")
        if hits:
            print(f"  row {r}: {hits}")
    # Also: dump rows 1-6 nonempty cells for top columns
    print("  first 6 rows (cols A..AE):")
    for r in range(1, 7):
        cells = []
        for c in range(1, 32):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                cells.append(f"{col_index_to_letter(c)}{r}={v!r}")
        print(f"    r{r}: {cells[:10]}")
    print()
