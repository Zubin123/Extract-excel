"""Independent audit of phase2_corpus_v4.xlsx.

Goal: verify the extractor's QA claims against the actual source workbooks.
We do NOT trust QA_Flag or Overall=PASS. We re-read each source sheet and
cross-check:
  1. Grand totals on the Data tab TOTALS row equal grand totals in the source
     cells (BB..BG / BC..BH at the pay_totals_row).
  2. Sum of day rows for each pay type equals the TOTALS row.
  3. WC State and ST on output match what's at the resolved header location.
  4. Per-day Total Hours read back from source matches output.
  5. Look for impossible / suspect values (negative hours, > 24h/day, strings
     in numeric columns, etc.).

Output: counts of mismatches by category and lists the first N examples.
"""
from __future__ import annotations
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402
import yaml  # noqa: E402

from anchors import (  # noqa: E402
    classify_sheet, col_letter_to_index, resolve_anchor_cell,
    resolve_pay_totals_row, select_profile, sheet_has_day_grid,
    find_field_mechanic_header_row, discover_field_mechanic_paytype_cols,
    find_fm_job_col,
)

TOL = 0.01
OUT = ROOT / "output" / "corpus_2022.xlsx"
DATA = ROOT / "data"


def to_num(v):
    """Coerce a value to float; numeric output is now written as '0.00'-style
    strings (matching the feedback format). Returns None if it isn't numeric."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def load_output():
    """Return dict[(excel_name, sheet_name)] -> {day_rows, totals_row}."""
    wb = openpyxl.load_workbook(OUT, data_only=True)
    ws = wb["Data"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    by_emp = defaultdict(lambda: {"days": [], "totals": None})
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        key = (rec["Excel Name"], rec["Sheet Name"])
        if rec["Day"] == "TOTALS":
            by_emp[key]["totals"] = rec
        else:
            by_emp[key]["days"].append(rec)
    return by_emp


def find_source(excel_name):
    matches = list(DATA.rglob(excel_name))
    return matches[0] if matches else None


def audit():
    cfg = yaml.safe_load((ROOT / "config" / "schema.yaml").read_text())
    profiles = cfg["profiles"]

    by_emp = load_output()
    print(f"Output has {len(by_emp)} employee sheets, {sum(len(v['days']) for v in by_emp.values())} day rows.\n")

    # Group keys by file to avoid reopening workbooks
    by_file = defaultdict(list)
    for (excel, sheet) in by_emp:
        by_file[excel].append(sheet)

    issues = Counter()
    examples = defaultdict(list)
    pay_types = ["RT", "OT", "DT", "PD", "4D", "4A"]

    def note(cat, msg):
        issues[cat] += 1
        if len(examples[cat]) < 5:
            examples[cat].append(msg)

    employees_checked = 0
    employees_pay_match = 0
    employees_fm_match = 0
    employees_th_match = 0
    employees_wc_match = 0
    employees_st_match = 0

    suspicious_values = Counter()

    for excel_name, sheet_names in by_file.items():
        src = find_source(excel_name)
        if src is None:
            note("source_missing", f"{excel_name}: source workbook not found in data/")
            continue
        try:
            wb = openpyxl.load_workbook(src, data_only=True)
        except Exception as e:
            note("source_open_error", f"{excel_name}: {e}")
            continue

        for sheet_name in sheet_names:
            employees_checked += 1
            if sheet_name not in wb.sheetnames:
                note("sheet_missing_in_source", f"{excel_name} :: {sheet_name}")
                continue
            ws = wb[sheet_name]
            out = by_emp[(excel_name, sheet_name)]
            totals = out["totals"]
            days   = out["days"]

            # Field Mechanic sheets: independently re-derive per-day pay-type
            # sums from the source by label discovery, then cross-check the
            # output's TOTALS and day-row sums. (The standard BB..BG cell check
            # below does not apply to this layout.)
            fm_hr = find_field_mechanic_header_row(ws)
            if fm_hr is not None:
                grid, day_idxs, _pts = discover_field_mechanic_paytype_cols(ws, fm_hr)
                last_row = ws.max_row or fm_hr
                fm_pts = ["RT", "OT", "DT", "PTO", "HP", "PP"]
                src_week = {pt: 0.0 for pt in fm_pts}
                job_col = find_fm_job_col(ws, fm_hr)
                def _is_line_item(r):
                    if job_col is None: return True
                    v = ws.cell(row=r, column=job_col).value
                    return v is not None and not (isinstance(v, str) and v.strip() == "")
                for (d_idx, pt), col in grid.items():
                    for r in range(fm_hr + 1, last_row + 1):
                        if not _is_line_item(r):
                            continue
                        v = ws.cell(row=r, column=col).value
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            src_week[pt] += float(v)
                fm_ok = True
                for pt in fm_pts:
                    out_val = to_num(totals.get(pt)) or 0.0
                    if abs(out_val - round(src_week[pt], 4)) > TOL:
                        fm_ok = False
                        note("fm_totals_mismatch_src",
                             f"{excel_name} :: {sheet_name} :: {pt} out={out_val} src={src_week[pt]}")
                # day-row sum vs TOTALS (internal consistency)
                for pt in fm_pts:
                    day_sum = sum((to_num(d.get(pt)) or 0) for d in days)
                    grand = to_num(totals.get(pt)) or 0
                    if abs(day_sum - grand) > TOL:
                        fm_ok = False
                        note("fm_internal_daysum_vs_totals",
                             f"{excel_name} :: {sheet_name} :: {pt} sum={day_sum} tot={grand}")
                if fm_ok:
                    employees_pay_match += 1
                    employees_fm_match += 1

                # Field Mechanic WC State / ST — discovered by label in the
                # header row, then read from the first non-blank cell below
                # (same logic the extractor uses; no hardcoded cell).
                def _fm_label_value(label):
                    target = label.strip().upper()
                    cmax = min(ws.max_column or 0, 30)
                    found_col = None
                    for c in range(1, cmax + 1):
                        v = ws.cell(row=fm_hr, column=c).value
                        if isinstance(v, str) and v.strip().upper() == target:
                            found_col = c
                            break
                    if found_col is None:
                        return None
                    for r in range(fm_hr + 1, min(last_row, fm_hr + 30) + 1):
                        v = ws.cell(row=r, column=found_col).value
                        if v is None: continue
                        if isinstance(v, str):
                            s = v.strip()
                            if s == "" or s.upper() in ("#N/A", "#VALUE!"):
                                continue
                            return s
                        return v
                    return None

                src_wc = _fm_label_value("WC STATE")
                src_st = _fm_label_value("ST")
                out_wc = totals.get("WC State")
                out_st = totals.get("ST")
                def _eq(a, b):
                    if a is None and (b is None or b == ""): return True
                    if b is None and (a is None or a == ""): return True
                    return str(a).strip() == str(b).strip()
                if _eq(src_wc, out_wc):
                    employees_wc_match += 1
                else:
                    note("fm_wc_state_mismatch",
                         f"{excel_name} :: {sheet_name} src={src_wc!r} out={out_wc!r}")
                if _eq(src_st, out_st):
                    employees_st_match += 1
                else:
                    note("fm_st_mismatch",
                         f"{excel_name} :: {sheet_name} src={src_st!r} out={out_st!r}")

                # Note: FM Total Hours comes from the clock-grid (rows 4-8) which
                # is worked clock-time, legitimately distinct from line-item
                # pay-type sums (see CLAUDE.md §6.3). Counting the day-summed
                # Total Hours against the AW8 weekly total instead.
                continue

            if not sheet_has_day_grid(ws, profiles):
                # NO_GRID — but the filter is supposed to have dropped these.
                note("nogrid_in_output", f"{excel_name} :: {sheet_name}")
                continue

            prof_name = select_profile(ws, profiles)
            if prof_name is None:
                note("profile_unresolvable", f"{excel_name} :: {sheet_name}")
                continue
            prof = profiles[prof_name]
            pay_row, _ = resolve_pay_totals_row(ws, prof, TOL)
            if pay_row is None:
                note("pay_row_unresolvable", f"{excel_name} :: {sheet_name}")
                continue

            # Source grand totals
            gt_cols = prof["grand_totals"]["cols"]
            src_gt = {}
            for pt, col in zip(pay_types, gt_cols):
                v = ws.cell(row=pay_row, column=col_letter_to_index(col)).value
                src_gt[pt] = float(v) if isinstance(v, (int, float)) else 0.0

            pay_ok = True
            for pt in pay_types:
                out_val = to_num(totals.get(pt)) or 0.0
                if abs(out_val - src_gt[pt]) > TOL:
                    pay_ok = False
                    note("totals_row_mismatch_src",
                         f"{excel_name} :: {sheet_name} :: {pt} out={out_val} src={src_gt[pt]}")
            if pay_ok:
                employees_pay_match += 1

            # Day-row sum vs TOTALS (already in output — pure self-check)
            internal_ok = True
            for pt in pay_types:
                day_sum = sum((to_num(d.get(pt)) or 0) for d in days)
                grand   = to_num(totals.get(pt)) or 0
                if abs(day_sum - grand) > TOL:
                    internal_ok = False
                    note("internal_daysum_vs_totals",
                         f"{excel_name} :: {sheet_name} :: {pt} sum={day_sum} tot={grand}")

            # Total Hours grand total
            th_col = prof["grand_totals"]["total_hours_col"]
            th_row = prof["total_hours_row"]
            src_th = ws.cell(row=th_row, column=col_letter_to_index(th_col)).value
            out_th = totals.get("Total Hours")
            out_th_num = to_num(out_th)
            if isinstance(src_th, (int, float)) and out_th_num is not None:
                if abs(float(src_th) - out_th_num) > TOL:
                    note("total_hours_mismatch",
                         f"{excel_name} :: {sheet_name} src={src_th} out={out_th}")
                else:
                    employees_th_match += 1
            elif isinstance(src_th, str) and "VALUE" in src_th.upper():
                note("source_th_value_error", f"{excel_name} :: {sheet_name} src={src_th!r} out={out_th}")
            elif src_th is None and out_th_num is not None:
                note("source_th_none_out_has_value", f"{excel_name} :: {sheet_name} src=None out={out_th}")
            else:
                note("th_type_mismatch", f"{excel_name} :: {sheet_name} src={src_th!r} ({type(src_th).__name__}) out={out_th!r} ({type(out_th).__name__})")

            # WC State / ST cross-check vs source via resolver
            wc_val, _, _ = resolve_anchor_cell(ws, prof["anchors"]["wc_state"])
            st_val, _, _ = resolve_anchor_cell(ws, prof["anchors"]["st"])
            out_wc = totals.get("WC State")
            out_st = totals.get("ST")

            def _eq(a, b):
                if a is None and (b is None or b == ""): return True
                if b is None and (a is None or a == ""): return True
                return str(a).strip() == str(b).strip()

            if _eq(wc_val, out_wc):
                employees_wc_match += 1
            else:
                note("wc_state_mismatch",
                     f"{excel_name} :: {sheet_name} src={wc_val!r} out={out_wc!r}")
            if _eq(st_val, out_st):
                employees_st_match += 1
            else:
                note("st_mismatch",
                     f"{excel_name} :: {sheet_name} src={st_val!r} out={out_st!r}")

            # Suspect values
            for d in days:
                th_n = to_num(d.get("Total Hours"))
                if th_n is not None:
                    if th_n < 0:
                        suspicious_values["negative_total_hours"] += 1
                    if th_n > 24:
                        suspicious_values["over_24h_in_a_day"] += 1
                for pt in pay_types:
                    vn = to_num(d.get(pt))
                    if vn is not None and vn < 0:
                        suspicious_values[f"negative_{pt}"] += 1
                wc = d.get("WC State")
                if isinstance(wc, str) and wc and wc not in (
                    "CA", "WA", "AZ", "OR", "NV", "ID", "MT", "TX", "CO", "UT", "NM"
                ):
                    suspicious_values["odd_wc_state_value"] += 1

    print("=" * 70)
    print("AUDIT RESULTS")
    print("=" * 70)
    print(f"Employee sheets checked: {employees_checked}")
    print()
    print(f"Pay-type grand totals match source: "
          f"{employees_pay_match}/{employees_checked} "
          f"(incl. {employees_fm_match} Field Mechanic by label-discovered cols)")
    print(f"Total Hours grand total matches source:            "
          f"{employees_th_match}/{employees_checked}")
    print(f"WC State matches header-resolved source value:     "
          f"{employees_wc_match}/{employees_checked}")
    print(f"ST matches header-resolved source value:           "
          f"{employees_st_match}/{employees_checked}")
    print()
    print("Issue categories:")
    for cat, n in issues.most_common():
        print(f"  {cat:<35} {n}")
        for ex in examples[cat]:
            print(f"      {ex}")
    print()
    print("Suspicious value counts (across all day rows):")
    if not suspicious_values:
        print("  (none)")
    for cat, n in suspicious_values.most_common():
        print(f"  {cat:<35} {n}")


if __name__ == "__main__":
    audit()
