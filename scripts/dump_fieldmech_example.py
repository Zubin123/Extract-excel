"""Dump a full Field Mechanic example sheet, cell by cell, with a side-by-side
comparison to a Standard sheet, so a human can open both in Excel and verify.
"""
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from anchors import col_index_to_letter

FM_FILE = ROOT / "data" / "sample1" / "WE 12 25 21 Chris Coulson 4916-6366-6087_1.xlsx"
FM_SHEET = "Job"

def dump_region(ws, label, rows, cols):
    print(f"\n   {label}")
    # header col letters
    hdr = "        " + "".join(f"{col_index_to_letter(c):>12}" for c in cols)
    print(hdr)
    for r in rows:
        line = f"   r{r:<4} "
        for c in cols:
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v)
            if len(s) > 11:
                s = s[:10] + "…"
            line += f"{s:>12}"
        print(line)


def main():
    wb = openpyxl.load_workbook(FM_FILE, data_only=True)
    ws = wb[FM_SHEET]
    print("=" * 78)
    print(f"FIELD MECHANIC EXAMPLE")
    print(f"File : {FM_FILE.name}")
    print(f"Sheet: {FM_SHEET!r}")
    print("=" * 78)

    print("\n-- Identity / title cells --")
    for addr in ("A1", "Z1", "A2", "B2", "D2", "A3", "B3"):
        print(f"   {addr:<4} = {ws[addr].value!r}")

    # Day labels are in row 3 at cols N S X AC AH AM AR
    day_cols = [14, 19, 24, 29, 34, 39, 44]  # N S X AC AH AM AR
    print("\n-- Day-label row (row 3) and date row (row 2) --")
    dump_region(ws, "rows 2-3 across day columns", [2, 3], day_cols)

    print("\n-- Time-label column (M = col 13) and time data, rows 4-8 --")
    dump_region(ws, "M-label + first 3 day blocks", [4, 5, 6, 7, 8],
                [13] + list(range(14, 30)))

    # Find where pay-type numbers / totals live: scan rows 8-45 in the first
    # day block and the far-right grand-total area.
    print("\n-- Scan for pay-type / totals rows (rows 8-45), first day block N..S --")
    dump_region(ws, "first day block cols N..S (14..19)", list(range(8, 46)),
                list(range(14, 20)))

    print("\n-- Far-right columns (potential grand totals), rows 8-45 --")
    # show columns AX..BL (50..64) to locate grand totals
    dump_region(ws, "cols AX..BL", list(range(8, 46)), list(range(50, 65)))


if __name__ == "__main__":
    main()
