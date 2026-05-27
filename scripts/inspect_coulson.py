"""Inspect Chris Coulson WE 12 25 21 to find the actual MON column."""
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl

from anchors import find_day_label_row, col_index_to_letter

path = ROOT / "data" / "sample1" / "WE 12 25 21 Chris Coulson 4916-6366-6087_1.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print(f"File: {path.name}")
print(f"Sheets: {wb.sheetnames}\n")

for sn in wb.sheetnames:
    ws = wb[sn]
    a1 = ws["A1"].value
    b2 = ws["B2"].value
    b1 = ws["B1"].value
    c2 = ws["C2"].value
    print(f"--- {sn!r} ---")
    print(f"  A1={a1!r}  B1={b1!r}  B2={b2!r}  C2={c2!r}")
    day = find_day_label_row(ws)
    if day is None:
        print("  no day-label row found in first 20 rows")
    else:
        row, col_idx = day
        print(f"  MON-row found at row={row}, first day col={col_index_to_letter(col_idx)} (col_idx={col_idx})")
    # Show row 4 contents up to column 30
    r4 = []
    for c in range(1, 35):
        v = ws.cell(row=4, column=c).value
        if v is not None:
            r4.append(f"{col_index_to_letter(c)}4={v!r}")
    print(f"  row4 nonempty: {r4[:15]}")
    print()
