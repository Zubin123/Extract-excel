"""Analyze phase2_corpus_v4 - Feedback.xlsx — find rows where the TRUE/FALSE
match grid shows FALSE, and the empty-employee sheets the user listed.

The feedback file structure:
  cols 1-25: our v4 Data row (with QA_Flag, Comments)
  col 26: blank separator
  cols 27-48: expected row
  col 49: blank separator
  cols 50+: per-field TRUE/FALSE comparison
"""
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
fb = openpyxl.load_workbook(ROOT / "phase2_corpus_v4 - Feedback.xlsx", data_only=True)
ws = fb["Data"]

# Headers
hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
# Block 1 = ours (1..25), expected = 27..48, match flags = 50..63
match_field_names = [hdrs[c-1] for c in range(50, 64)]
print("Match-flag columns:", match_field_names)

# Build maps: row -> our row dict, expected row dict, match dict
total = 0
all_match = 0
mismatches_by_field = Counter()
mismatch_rows = []
empty_rows_in_ours = []  # rows where our Date/Start/Stop are all None but expected has values

EMPTY_EMP_SHEETS = {
    ("WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx", "Job"),
    ("WE 12 25 21 Hall 4906-2066-0647_1.xlsx", "JOB"),
    ("WE 12 25 21 Juan Medina 4920-9705-5655_1.xlsx", "Job"),
    ("WE 12 25 21 Michael Beeler 4902-6840-4647_1.xlsx", "Job"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "JOB"),
    ("WE 12 25 21 Smith, Dustin 4922-4811-6135_1.xlsx", "OVERHEAD - CA"),
    ("WE 12 25 21 6562 UPL -WA 4937-5257-4377_1.xlsx", "Wilson G."),
    ("WE 12 25 21 6562 UPL -WA 4937-5257-4377_1.xlsx", "Jackson HP-PTO"),
    ("WE 12 25 21 6799 McIntire UPL -WA 4919-9109-7769_1.xlsx", "McIntire"),
    ("WE 12 25 21 6818 UPL -WA 4903-4699-6137_1.xlsx", "I Christopherson PTO-HP"),
    ("WE 12 25 21 CA 6721 UPL -WA 4921-9281-7577_1.xlsx", "David Cortez PTO-HP"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Case"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Sanders"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Eric Schmidt"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Raper  HP-PTO"),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Peake "),
    ("WE 12 25 21 CA GF'S UPL -WA 4924-2782-9673_1.xlsx", "Zeulner"),
}

empty_sheet_rows = defaultdict(list)

for r in range(2, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    our_file = row[2]
    our_sheet = row[3]
    our_day = row[9]
    our_date = row[8]
    our_start = row[10]
    our_stop = row[13]
    our_th = row[14]

    expected_date = row[55] if len(row) > 55 else None
    expected_start = row[57] if len(row) > 57 else None

    # Collect rows for the listed empty-employee sheets
    key = (our_file, our_sheet)
    if key in EMPTY_EMP_SHEETS:
        empty_sheet_rows[key].append({
            "Day": our_day, "Date": our_date,
            "Start": our_start, "Stop": our_stop, "TH": our_th
        })

    # Per-field match grid (cols 50..63)
    flags = [ws.cell(row=r, column=c).value for c in range(50, 64)]
    total += 1
    if all(f is True or f == "TRUE" for f in flags if f is not None):
        all_match += 1
    else:
        for fn, fv in zip(match_field_names, flags):
            if fv is False or fv == "FALSE":
                mismatches_by_field[fn] += 1
        mismatch_rows.append((r, our_file, our_sheet, our_day,
                               [fn for fn, fv in zip(match_field_names, flags)
                                if fv is False or fv == "FALSE"]))

print(f"\nTotal data rows in feedback file: {total}")
print(f"Rows with all fields matching:    {all_match}")
print(f"Rows with at least one mismatch:  {total - all_match}")
print(f"\nMismatch counts per field (from feedback file's own TRUE/FALSE grid):")
for fn, n in mismatches_by_field.most_common():
    print(f"  {fn:<20} {n}")

print(f"\n=== Empty-employee sheets the user listed ===")
print(f"Found {len(empty_sheet_rows)}/{len(EMPTY_EMP_SHEETS)} listed sheets in feedback file:")
for key, rows in sorted(empty_sheet_rows.items()):
    has_data = any(
        r["Start"] not in (None, "") or r["Stop"] not in (None, "") or
        (isinstance(r["TH"], (int, float)) and r["TH"] > 0) or
        (isinstance(r["TH"], str) and r["TH"] not in ("", "0.00"))
        for r in rows if r["Day"] != "TOTALS"
    )
    th_grand = next((r["TH"] for r in rows if r["Day"] == "TOTALS"), None)
    print(f"  {key[0]} :: {key[1]!r}")
    print(f"    rows: {len(rows)}  has-day-data: {has_data}  TOTALS Total Hours: {th_grand!r}")

print(f"\n=== Sample of first 10 mismatch rows ===")
for r, f, s, d, fields in mismatch_rows[:10]:
    print(f"  row {r}: {f[:40]:<40} {s[:15]:<15} {d:<6} mismatched: {fields}")
