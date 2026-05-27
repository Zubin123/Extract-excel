"""Search the WE 12 18 21 CA 6427 workbook for EE ID 3398 (Delgado) anywhere."""
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
DATA = ROOT / "data"

xlsx = next(DATA.rglob("*CA 6427 - UPLOADED*4901*.xlsx"))
wb = load_workbook(xlsx, data_only=True)

print(f"File: {xlsx.name}")
print(f"Sheets: {wb.sheetnames}\n")

# Look at first few rows of the Delgado sheet to understand structure
ws = wb["Delgado"]
print("DELGADO sheet — rows 1-15, cols A-J:")
for r in range(1, 16):
    line = []
    for c in range(1, 11):
        v = ws.cell(row=r, column=c).value
        col_letter = chr(ord("A") + c - 1)
        if v is not None:
            line.append(f"{col_letter}{r}={v!r}")
    if line:
        print(f"  row {r}: " + " | ".join(line))

# Check the Employees sheet — look at ID ranges
emp_ws = wb["Employees"]
print(f"\nEmployees sheet has {emp_ws.max_row} rows")
print("Looking for ID 3398 anywhere in column A or B:")
for r in range(1, emp_ws.max_row + 1):
    a = emp_ws.cell(row=r, column=1).value
    b = emp_ws.cell(row=r, column=2).value
    c = emp_ws.cell(row=r, column=3).value
    if str(a).strip() == "3398" or b == 3398:
        print(f"  row {r}: A={a!r}, B={b!r}, C={c!r}")

# Show min/max ID range
ids = []
for r in range(3, emp_ws.max_row + 1):
    b = emp_ws.cell(row=r, column=2).value
    if isinstance(b, int):
        ids.append(b)
print(f"\nEmployees roster: {len(ids)} IDs ranging {min(ids)}-{max(ids)}")
print(f"Sample IDs around 3398: {sorted(i for i in ids if 3300 <= i <= 3500)[:20]}")

# Maybe 3398 is in some other column?
print("\nSearching ALL cells in Employees sheet for value 3398:")
hits = 0
for r in range(1, emp_ws.max_row + 1):
    for c in range(1, emp_ws.max_column + 1):
        v = emp_ws.cell(row=r, column=c).value
        if v == 3398 or str(v).strip() == "3398":
            col_letter = chr(ord("A") + c - 1) if c <= 26 else "AA"
            print(f"  Found at {col_letter}{r}: {v!r}")
            hits += 1
if not hits:
    print("  not found")

# Check the 'Cline' sheet (which we extract correctly) for reference structure
print("\nCLINE sheet (a working sheet) — rows 1-3:")
ws = wb["Cline"]
for r in range(1, 4):
    line = []
    for c in range(1, 11):
        v = ws.cell(row=r, column=c).value
        col_letter = chr(ord("A") + c - 1)
        if v is not None:
            line.append(f"{col_letter}{r}={v!r}")
    if line:
        print(f"  row {r}: " + " | ".join(line))

wb.close()
