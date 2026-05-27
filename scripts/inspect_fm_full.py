"""Full raw-grid dump of Field Mechanic sheets — assumption-free.

Goal: SEE the real structure before encoding any discovery logic. We do NOT
assume pay_block_width, pay-type order, grand-total columns, or the pay-totals
row. We print the raw cell grid and let the structure reveal itself.

For each Field Mechanic candidate sheet we print:
  - identity cells (A1,B1,B2,C2)
  - the day-label row (full-name tokens) and exact columns
  - every non-empty cell in rows 1..(day_row+12), col 1..80, as a compact grid
  - rows that look numeric-heavy (candidate pay/totals rows)

Run: .venv\\Scripts\\python.exe scripts\\inspect_fm_full.py
"""
from __future__ import annotations
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from datetime import datetime, date, time
from anchors import col_index_to_letter

DAY_FULL = {"MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"}
DAY_SHORT = {"MON", "TUE", "WED", "THUR", "THU", "FRI", "SAT", "SUN"}

# Representative Field Mechanic files (from SESSION_HANDOFF §4 census).
TARGETS = [
    ("data/sample1/WE 12 25 21 Chris Coulson 4916-6366-6087_1.xlsx", None),
    ("data/sample1/WE 12 25 21 Ed Pottage 4929-5262-8135_1.xlsx", None),
    ("data/sample1/WE 12 25 21 Hall 4906-2066-0647_1.xlsx", None),
]


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date, time)):
        return f"<{type(v).__name__}:{v}>"
    if isinstance(v, float):
        return f"{v:g}"
    if isinstance(v, str):
        s = v.strip()
        return s[:14]
    return str(v)


def find_day_rows(ws, max_r=20, max_c=80):
    """Return list of (row, [cols], style) where style in {full,short}."""
    out = []
    for r in range(1, max_r + 1):
        full_cols, short_cols = [], []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u in DAY_FULL:
                    full_cols.append(c)
                elif u in DAY_SHORT:
                    short_cols.append(c)
        if len(full_cols) >= 3:
            out.append((r, full_cols, "full"))
        elif len(short_cols) >= 3:
            out.append((r, short_cols, "short"))
    return out


def dump_grid(ws, r0, r1, c0, c1):
    for r in range(r0, r1 + 1):
        cells = []
        for c in range(c0, c1 + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                cells.append(f"{col_index_to_letter(c)}{r}={fmt(v)}")
        if cells:
            print(f"    r{r:>2}: " + "  ".join(cells))


def main():
    for rel, _ in TARGETS:
        path = ROOT / rel
        if not path.exists():
            print(f"!! MISSING: {rel}\n")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        print("=" * 100)
        print(f"FILE: {path.name}")
        print(f"Sheets: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            a1, b1, b2, c2 = (ws["A1"].value, ws["B1"].value,
                              ws["B2"].value, ws["C2"].value)
            day_rows = find_day_rows(ws)
            if not day_rows:
                continue  # not a grid sheet; skip in this probe
            print("-" * 100)
            print(f"SHEET {sn!r}")
            print(f"  A1={a1!r} B1={b1!r} B2={b2!r} C2={c2!r}")
            for (r, cols, style) in day_rows:
                strides = [cols[i + 1] - cols[i] for i in range(len(cols) - 1)]
                letters = [col_index_to_letter(c) for c in cols]
                print(f"  DAY ROW {r} ({style}): cols={letters} strides={strides}")
            # Dump full grid from row 1 down to a bit past the last day row.
            last_day = max(r for (r, _, _) in day_rows)
            c_max = min(ws.max_column or 0, 80)
            print(f"  --- raw grid rows 1..{last_day + 14}, cols A..{col_index_to_letter(c_max)} ---")
            dump_grid(ws, 1, last_day + 14, 1, c_max)
            print()


if __name__ == "__main__":
    main()
