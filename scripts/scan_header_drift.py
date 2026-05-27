"""Scan all employee sheets in the corpus for header-row drift.

For each employee sheet, read row 12 columns A..L (where the field labels
live) and tally how often each header pattern appears. Any sheet whose
G12/H12/I12 don't match the assumed "WC STATE"/"ST" layout indicates a
file where wc_state and st are being misread.

Also checks:
  - column A labels for rows 5-8 (should be Start / Lunch Out / Lunch In / Stop)
  - row that says RT/OT/DT/PD/4D/4A pay-type labels (if any)
"""
from collections import defaultdict
from openpyxl import load_workbook
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

ROOT = Path(r"c:\Users\MohammedZubinEssudee\OneDrive - iBridge Global Services\Desktop\Extract Excel")
DATA = ROOT / "data"

# Pattern bucket: row 12 headers tuple -> list of (file, sheet) that have it
patterns = defaultdict(list)
time_label_patterns = defaultdict(list)
pay_label_patterns = defaultdict(list)

# Per-file impact tally
wc_misread_count = 0
st_misread_count = 0
total_employee_sheets = 0


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if isinstance(v, str):
        return v.strip().upper()
    return v


for xlsx in sorted(DATA.rglob("*.xlsx")):
    if "Header Template" in xlsx.name:
        continue
    try:
        wb = load_workbook(xlsx, data_only=True, read_only=False)
    except Exception:
        continue
    for ws in wb.worksheets:
        a1 = ws["A1"].value
        b2 = ws["B2"].value
        is_real_employee_sheet = (
            isinstance(a1, str) and (";" in a1 or "," in a1)
            and isinstance(b2, int)
        )
        if not is_real_employee_sheet:
            continue
        total_employee_sheets += 1

        # Row 12 header pattern (cols A..L = 1..12)
        row12 = tuple(cell(ws, 12, c) for c in range(1, 13))
        patterns[row12].append((xlsx.name, ws.title))

        # Where is "WC STATE" actually?
        wc_state_col = None
        st_col = None
        for c in range(1, 15):
            v = cell(ws, 12, c)
            if v == "WC STATE":
                wc_state_col = c
            if v == "ST":
                st_col = c
        # Extractor reads G13 (col 7) for wc_state and H13 (col 8) for st
        if wc_state_col is not None and wc_state_col != 7:
            wc_misread_count += 1
        if st_col is not None and st_col != 8:
            st_misread_count += 1

        # Time-row labels: column A or B for rows 5-8
        # The labels are usually in row 4 column A or just near rows 5-8
        # Quick check: read col A rows 5-8
        time_labels = tuple(cell(ws, r, 1) for r in (5, 6, 7, 8))
        # Many sheets have empty col A here; only record patterns where any non-None
        if any(time_labels):
            time_label_patterns[time_labels].append((xlsx.name, ws.title))

        # Pay-type labels typically in row above day grid; look for RT/OT/DT/PD/4D/4A
        # Day labels are in row 4; pay-type sub-labels often in row 4 or row 5 at
        # the start of each day block (e.g., L4=MON, L5=RT etc.). Try row 4 with offset:
        # MON typically at col 12 (L) for standard. Pull next 6 cells at row 4 starting L.
        # Actually the pay-type labels often live in row 4 just below MON, or in row 3.
        # Look at row 4 cols L..Q to see what's there.
        pass

    wb.close()

# Report row-12 header patterns
print(f"Total employee sheets scanned: {total_employee_sheets}")
print()
print("=" * 88)
print("ROW-12 HEADER PATTERNS (cols A..L), grouped by frequency")
print("=" * 88)
for headers, files in sorted(patterns.items(), key=lambda x: -len(x[1])):
    print(f"\n--- {len(files)} sheet(s) with pattern:")
    for i, h in enumerate(headers, start=1):
        col = chr(ord("A") + i - 1)
        # mark the WC STATE / ST cols
        marker = ""
        if h == "WC STATE":
            marker = "  <-- WC STATE at col " + col + (
                "  (extractor reads G — OK)" if col == "G"
                else f"  (extractor reads G — WILL BE WRONG, real value at {col})"
            )
        elif h == "ST":
            marker = "  <-- ST at col " + col + (
                "  (extractor reads H — OK)" if col == "H"
                else f"  (extractor reads H — WILL BE WRONG, real value at {col})"
            )
        print(f"    {col}12 = {h!r}{marker}")
    print(f"  Sample files (first 3):")
    for f, s in files[:3]:
        print(f"    {f}  |  sheet={s!r}")

print()
print("=" * 88)
print(f"IMPACT: {wc_misread_count}/{total_employee_sheets} sheets have WC State misread")
print(f"IMPACT: {st_misread_count}/{total_employee_sheets} sheets have ST misread")
print("=" * 88)
