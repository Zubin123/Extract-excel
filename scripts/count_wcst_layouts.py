"""Tally distinct on-sheet positions for WC State and ST across the corpus.

For every employee-style sheet in data/ we record:
  - which profile matched (standard / shifted)
  - where the WC STATE header label was actually found (row, col)
  - where the first non-empty cell below that header lives (data row)
  - same for ST

We print distinct (header_row, header_col, data_row) tuples and counts.
"""
from __future__ import annotations
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402
import yaml  # noqa: E402

from anchors import (  # noqa: E402
    classify_sheet, col_index_to_letter, col_letter_to_index,
    select_profile, sheet_has_day_grid,
)


def find_header_and_data(ws, label: str, expected_header_row: int,
                         expected_col_letter: str,
                         row_radius: int = 1, search_cols: int = 14,
                         data_depth: int = 5):
    """Return (header_row, header_col_letter, data_row, value) or None."""
    label_u = label.strip().upper()
    expected_col = col_letter_to_index(expected_col_letter)

    rows_to_try = [expected_header_row]
    for d in range(1, row_radius + 1):
        rows_to_try += [expected_header_row - d, expected_header_row + d]
    rows_to_try = [r for r in rows_to_try if r >= 1]

    max_c = min(ws.max_column or 0, search_cols)

    def _norm(v):
        return v.strip().upper() if isinstance(v, str) else None

    for hr in rows_to_try:
        found_col = None
        if _norm(ws.cell(row=hr, column=expected_col).value) == label_u:
            found_col = expected_col
        else:
            for c in range(1, max_c + 1):
                if _norm(ws.cell(row=hr, column=c).value) == label_u:
                    found_col = c
                    break
        if found_col is None:
            continue
        for r in range(hr + 1, hr + 1 + data_depth):
            v = ws.cell(row=r, column=found_col).value
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                return hr, col_index_to_letter(found_col), r, v
        return hr, col_index_to_letter(found_col), None, None
    return None


def main():
    cfg = yaml.safe_load((ROOT / "config" / "schema.yaml").read_text())
    profiles = cfg["profiles"]

    files = sorted((ROOT / "data").rglob("*.xlsx"))
    files = [p for p in files if not p.name.startswith("~$") and "Header Template" not in p.name]
    print(f"Scanning {len(files)} workbooks...\n")

    wc_positions = Counter()   # (header_row, header_col, data_row)
    st_positions = Counter()
    wc_cells = Counter()       # final cell address used
    st_cells = Counter()
    profile_counts = Counter()
    sheets_scanned = 0
    sheets_skipped_no_grid = 0

    wc_missing = 0
    st_missing = 0

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            print(f"  ! skip {f.name}: {e}")
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            kind, _ = classify_sheet(ws, profiles)
            if kind in ("reference", "unfilled_template"):
                continue
            # Only profile-matching grid sheets have meaningful anchor positions
            if not sheet_has_day_grid(ws, profiles):
                sheets_skipped_no_grid += 1
                continue
            prof_name = select_profile(ws, profiles)
            if prof_name is None:
                continue
            prof = profiles[prof_name]
            profile_counts[prof_name] += 1
            sheets_scanned += 1

            wc_spec = prof["anchors"]["wc_state"]
            st_spec = prof["anchors"]["st"]

            wc = find_header_and_data(ws, wc_spec["header_label"],
                                       wc_spec["header_row"],
                                       "".join(c for c in wc_spec["cell"] if c.isalpha()))
            st = find_header_and_data(ws, st_spec["header_label"],
                                       st_spec["header_row"],
                                       "".join(c for c in st_spec["cell"] if c.isalpha()))

            if wc is None:
                wc_missing += 1
            else:
                hr, hc, dr, _v = wc
                wc_positions[(hr, hc, dr)] += 1
                if dr is not None:
                    wc_cells[f"{hc}{dr}"] += 1

            if st is None:
                st_missing += 1
            else:
                hr, hc, dr, _v = st
                st_positions[(hr, hc, dr)] += 1
                if dr is not None:
                    st_cells[f"{hc}{dr}"] += 1

    print(f"Grid-bearing employee sheets scanned: {sheets_scanned}")
    print(f"Profiles:")
    for p, n in profile_counts.most_common():
        print(f"  {p:<10} {n}")
    print()

    def dump(label, positions, cells, missing):
        print(f"== {label} ==")
        print(f"  missing (label not found): {missing}")
        print(f"  distinct (header_row, header_col, data_row) tuples: {len(positions)}")
        for (hr, hc, dr), n in positions.most_common():
            print(f"    header={hc}{hr}  data_row={dr}   sheets={n}")
        print(f"  distinct final data-cell addresses: {len(cells)}")
        for addr, n in cells.most_common():
            print(f"    {addr:<6} {n}")
        print()

    dump("WC State", wc_positions, wc_cells, wc_missing)
    dump("ST",       st_positions, st_cells, st_missing)


if __name__ == "__main__":
    main()
