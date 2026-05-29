"""Compare the manager's expected output (v4 feedback file) against the
current phase2_corpus_v4.xlsx, field by field.

The feedback file has THREE blocks on the Data tab (per SESSION_HANDOFF §6):
  - cols 1-24  : our old v4 snapshot at the time the feedback was authored
  - col 25     : 'Comments' (manager annotations)
  - cols 27-48 : EXPECTED output (the manager's truth)
  - cols 50-63 : per-field TRUE/FALSE match grid

We deliberately ignore the stale cols-1-24 snapshot. We compare cols 27-48
(EXPECTED) against the CURRENT output, joined on (Excel Name, Sheet Name, Day).
This is the only honest measure — it answers "does our output now match what
the manager wanted?"
"""
from __future__ import annotations
import sys
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CURRENT = ROOT / "output" / "corpus_v5_final.xlsx"
FEEDBACK = ROOT / "phase2_corpus_v4 - Feedback.xlsx"

# Expected-block columns in the feedback file (cols 27..48, 1-based).
EXPECTED_COLS = {
    "Folder Name": 27,
    "Excel Name":  28,
    "Sheet Name":  29,
    "Employee Name": 30,
    "EE ID":       31,
    "WC State":    32,
    "ST":          33,
    # col 34 is a stray "f" header
    "Day":         35,
    "Start":       36,
    "Lunch Out":   37,
    "Lunch In":    38,
    "Stop":        39,
    "Total Hours": 40,
    "RT": 41, "OT": 42, "DT": 43, "PD": 44, "4D": 45, "4A": 46,
    "PTO": 47, "HP": 48,
}
COMPARE_FIELDS = [
    "Employee Name", "WC State", "ST",
    "Start", "Lunch Out", "Lunch In", "Stop", "Total Hours",
    "RT", "OT", "DT", "PD", "4D", "4A", "PTO", "HP",
]


def norm(v):
    """Normalize a value for comparison: None/blank-string/NaN -> None;
    numerics with float-equality; strings stripped."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s == "" else s
    if isinstance(v, float) and v != v:  # NaN
        return None
    return v


def numeric_close(a, b, tol=0.01):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def eq(a, b, field):
    a, b = norm(a), norm(b)
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # Numeric fields: tolerant float compare
    if field in {"Total Hours", "RT", "OT", "DT", "PD", "4D", "4A", "PTO", "HP"}:
        if numeric_close(a, b):
            return True
        # Could be that one side is "0" string and other is 0.0
    if field == "EE ID":
        try:
            return int(float(a)) == int(float(b))
        except (TypeError, ValueError):
            return str(a) == str(b)
    # Time strings or names: case-insensitive string compare with trimming
    return str(a).strip().lower() == str(b).strip().lower()


def load_expected():
    """Return dict[(excel, sheet, day)] -> {field: value}, plus comments list."""
    wb = openpyxl.load_workbook(FEEDBACK, data_only=True)
    ws = wb["Data"]
    rows = {}
    comments = []
    for r in range(2, (ws.max_row or 1) + 1):
        excel = ws.cell(row=r, column=EXPECTED_COLS["Excel Name"]).value
        sheet = ws.cell(row=r, column=EXPECTED_COLS["Sheet Name"]).value
        day   = ws.cell(row=r, column=EXPECTED_COLS["Day"]).value
        if not excel:
            continue
        key = (str(excel).strip(), str(sheet).strip(), str(day).strip() if day else "")
        rec = {}
        for field, col in EXPECTED_COLS.items():
            if field in ("Folder Name", "Excel Name", "Sheet Name", "Day"):
                continue
            rec[field] = ws.cell(row=r, column=col).value
        rows[key] = rec
        cmt = ws.cell(row=r, column=25).value
        if isinstance(cmt, str) and cmt.strip():
            comments.append((key, cmt.strip()))
    return rows, comments


def load_current():
    """Return dict[(excel, sheet, day)] -> {field: value}."""
    wb = openpyxl.load_workbook(CURRENT, data_only=True)
    ws = wb["Data"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        excel = row[idx["Excel Name"]]
        sheet = row[idx["Sheet Name"]]
        day   = row[idx["Day"]]
        if excel is None:
            continue
        key = (str(excel).strip(), str(sheet).strip(), str(day).strip() if day is not None else "")
        rows[key] = {f: row[idx[f]] for f in COMPARE_FIELDS if f in idx}
        rows[key]["EE ID"] = row[idx["EE ID"]] if "EE ID" in idx else None
    return rows


def main():
    print(f"Current output: {CURRENT.name}")
    print(f"Feedback file:  {FEEDBACK.name}")
    expected, comments = load_expected()
    current = load_current()
    print(f"\nExpected rows: {len(expected)}    Current rows: {len(current)}")

    # 1) Coverage
    expected_keys = set(expected)
    current_keys = set(current)
    missing_in_current = expected_keys - current_keys
    extra_in_current = current_keys - expected_keys
    common = expected_keys & current_keys
    print(f"\nCoverage:")
    print(f"  Rows in BOTH:                 {len(common)}")
    print(f"  Expected but missing now:     {len(missing_in_current)}")
    print(f"  Present now but not expected: {len(extra_in_current)}  (mostly Field Mechanic — new content)")
    for k in sorted(missing_in_current)[:8]:
        print(f"    MISSING: {k}")
    if len(missing_in_current) > 8:
        print(f"    ... and {len(missing_in_current)-8} more")

    # 2) Field-by-field match on common rows
    # Treat expected=None as "not specified" rather than "must be blank" — the
    # manager's expected block intentionally leaves pay-type cells blank on
    # TOTALS rows and on rows where the field doesn't apply. We only count
    # rows where the manager EXPLICITLY specified a value.
    field_pass = Counter()
    field_fail = Counter()
    field_total_specified = Counter()
    sample_fails = defaultdict(list)
    for key in common:
        e, c = expected[key], current[key]
        for f in COMPARE_FIELDS:
            ev, cv = e.get(f), c.get(f)
            if norm(ev) is None:
                continue  # manager didn't specify a value here — skip
            field_total_specified[f] += 1
            if eq(ev, cv, f):
                field_pass[f] += 1
            else:
                field_fail[f] += 1
                if len(sample_fails[f]) < 4:
                    sample_fails[f].append((key, ev, cv))

    print("\nField-by-field accuracy on rows where the manager SPECIFIED a value:")
    fields_sorted = sorted(set(field_pass) | set(field_fail))
    for f in fields_sorted:
        p, fl = field_pass[f], field_fail[f]
        tot = field_total_specified[f]
        pct = (100*p/tot) if tot else 100
        flag = "" if fl == 0 else f"   <-- {fl} mismatches"
        print(f"  {f:14}  {p:>5}/{tot:<5}  ({pct:5.1f}%){flag}")

    # 3) Sample failures
    print("\nSample mismatches (max 4 per field):")
    for f, samples in sample_fails.items():
        print(f"  [{f}]")
        for (key, ev, cv) in samples:
            print(f"    {key}  expected={ev!r}  got={cv!r}")

    # 4) Manager comments
    print(f"\nManager comments on feedback file: {len(comments)}")
    cmt_count = Counter(c for _,c in comments)
    for c, n in cmt_count.most_common(10):
        print(f"  {n:>4}  {c[:90]}")


if __name__ == "__main__":
    main()
