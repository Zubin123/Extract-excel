"""Exhaustive structural census of every employee sheet across the corpus.

For each sheet that looks like an employee sheet (by ANY of several
loose structural signals — not by our classifier), record:

  - Identity layout: where the name and EE ID live (A1/B2, B1/C2, etc.)
  - Day-label row + first day column + stride (gap between day columns)
  - Day-label token style (MON vs MONDAY)
  - Date row position
  - Time-label column and first time-data row
  - Pay-type column order in the day-block header (where present)
  - WC State / ST header position and data row
  - Pay-totals row (discovered via existing sum-cross-check)

Then aggregate and report distinct (cell-position) layouts seen.
"""
from __future__ import annotations
import sys
import re
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl

DAY_SHORT = {"MON","TUE","WED","THUR","THU","FRI","SAT","SUN"}
DAY_FULL  = {"MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY"}
TIME_LABELS = {"START","STOP","LUNCH OUT","LUNCH IN"}
PAY_TYPE_TOKENS = {"RT","OT","DT","PD","4D","4A"}
NAME_RE = re.compile(r"^[A-Z][a-zA-Z'\-]+\s*[;,]\s*[A-Z][a-zA-Z'\-]+", re.UNICODE)


def col_letter(idx):
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def is_int_like(v):
    if isinstance(v, bool): return False
    if isinstance(v, int): return True
    if isinstance(v, float): return v.is_integer()
    return False


def find_day_row(ws):
    """Find the row that has >=5 day-tokens. Return (row, [(col, token_style)])."""
    max_r = min(ws.max_row or 0, 30)
    max_c = min(ws.max_column or 0, 80)
    for r in range(1, max_r + 1):
        hits = []
        style = None
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u in DAY_SHORT:
                    hits.append((c, "short"))
                    style = "short"
                elif u in DAY_FULL:
                    hits.append((c, "full"))
                    style = "full"
        if len(hits) >= 5:
            return r, hits, style
    return None


def find_time_label_col(ws, day_row):
    """Find column where >=3 of {Start,Stop,Lunch In,Lunch Out} appear within
    rows day_row..day_row+8. Return (col, [(label, row)])."""
    max_c = min(ws.max_column or 0, 80)
    for c in range(1, max_c + 1):
        hits = []
        for r in range(day_row, day_row + 10):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() in TIME_LABELS:
                hits.append((v.strip().upper(), r))
        if len(hits) >= 3:
            return c, hits
    return None


def find_date_row(ws, day_row):
    """The row near day_row that has the most date-typed cells."""
    best = None
    for r in (day_row - 1, day_row - 2, day_row + 1, day_row + 2):
        if r < 1: continue
        n = 0
        for c in range(1, min(ws.max_column or 0, 80) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (datetime, date)):
                n += 1
        if n >= 5 and (best is None or n > best[1]):
            best = (r, n)
    return best[0] if best else None


def find_pay_type_header(ws, day_row, day_cols):
    """Look for pay-type tokens (RT/OT/...) in rows day_row..day_row+2 within
    the day-block columns. Return ordered list of (col_offset, token)."""
    if not day_cols: return []
    first_day_col = day_cols[0][0]
    stride = (day_cols[1][0] - day_cols[0][0]) if len(day_cols) >= 2 else 6
    for r in (day_row + 1, day_row + 2):
        hits = []
        for k in range(stride):
            v = ws.cell(row=r, column=first_day_col + k).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u in PAY_TYPE_TOKENS:
                    hits.append((k, u))
        if len(hits) >= 4:
            return hits, r, stride
    return [], None, stride


def find_wcst_positions(ws):
    """Scan rows 1-20 for the WC STATE and ST header text. Return positions."""
    wc, st = None, None
    max_r = min(ws.max_row or 0, 20)
    max_c = min(ws.max_column or 0, 20)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                u = v.strip().upper()
                if u == "WC STATE" and wc is None:
                    wc = (r, c)
                elif u == "ST" and st is None:
                    st = (r, c)
    return wc, st


def identity_layout(ws):
    """Where do name + EE ID live? Return (name_addr, id_addr) or None."""
    a1 = ws["A1"].value
    b1 = ws["B1"].value
    b2 = ws["B2"].value
    c2 = ws["C2"].value
    if isinstance(a1, str) and NAME_RE.match(a1.strip()) and is_int_like(b2):
        return ("A1", "B2", "standard")
    if (a1 is None or (isinstance(a1, str) and a1.strip() == "")) and \
       isinstance(b1, str) and NAME_RE.match(b1.strip()) and is_int_like(c2):
        return ("B1", "C2", "alt")
    if isinstance(a1, str) and a1.strip().upper().startswith("INPUT") and is_int_like(b2):
        return ("A1-placeholder", "B2", "placeholder")
    if isinstance(a1, str) and a1.strip().upper().startswith("INPUT") and b2 is None:
        return ("A1-placeholder", "B2-empty", "template-empty")
    if isinstance(b2, str) and ("EE" in b2.upper() and ("#" in b2 or ":" in b2)):
        return ("A1", "B2-EE-literal", "template-empty")
    return None


def looks_employee_like(ws):
    """Loose filter: sheet has either an identity layout OR a day-row.

    We deliberately use a loose definition so we count sheets our current
    classifier ignores (like the Field Mechanic sheets)."""
    if identity_layout(ws) is not None:
        return True
    if find_day_row(ws) is not None:
        return True
    return False


def is_reference_sheet(ws_name):
    n = ws_name.strip().upper()
    return n in {"JOB LIST","EQUIPMENT","EMPLOYEES","COLUMNLISTS","TEMPLATE",
                 "MASTER TEMPLATE","TRAINING"} or n.startswith("MASTER TEMPLATE")


def census():
    files = sorted((ROOT / "data").rglob("*.xlsx"))
    files = [p for p in files if not p.name.startswith("~$") and "Header Template" not in p.name]
    print(f"Scanning {len(files)} workbooks...\n")

    total_workbooks = len(files)
    total_sheets_seen = 0
    total_employee_like = 0
    total_reference_by_name = 0

    identity_layouts = Counter()
    day_row_positions = Counter()
    day_first_col_positions = Counter()
    day_stride = Counter()
    day_token_style = Counter()
    date_row_offsets = Counter()  # date_row - day_row
    time_label_cols = Counter()
    time_first_row_offsets = Counter()  # first time-label row - day_row
    pay_type_order = Counter()
    pay_type_header_offset = Counter()  # pay-type header row - day_row
    wc_positions = Counter()
    st_positions = Counter()
    profile_signatures = Counter()  # full structural signature

    # Per-template-family classification
    layout_family = Counter()

    file_details = []

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            print(f"  ! cannot open {f.name}: {e}")
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            total_sheets_seen += 1
            if is_reference_sheet(sn):
                total_reference_by_name += 1
                continue
            if not looks_employee_like(ws):
                total_reference_by_name += 1
                continue

            total_employee_like += 1

            ident = identity_layout(ws)
            identity_layouts[ident[2] if ident else "no-identity"] += 1

            day = find_day_row(ws)
            wc, st = find_wcst_positions(ws)
            if wc is not None:
                wc_positions[(wc[0], col_letter(wc[1]))] += 1
            if st is not None:
                st_positions[(st[0], col_letter(st[1]))] += 1

            if day is None:
                layout_family[("no-day-row", ident[2] if ident else None)] += 1
                continue

            day_row, day_cols, style = day
            first_col = day_cols[0][0]
            stride = (day_cols[1][0] - day_cols[0][0]) if len(day_cols) >= 2 else None

            day_row_positions[day_row] += 1
            day_first_col_positions[col_letter(first_col)] += 1
            day_stride[stride] += 1
            day_token_style[style] += 1

            date_r = find_date_row(ws, day_row)
            if date_r is not None:
                date_row_offsets[date_r - day_row] += 1

            tl = find_time_label_col(ws, day_row)
            if tl is not None:
                tl_col, tl_hits = tl
                time_label_cols[col_letter(tl_col)] += 1
                first_tl_row = min(r for (_lbl, r) in tl_hits)
                time_first_row_offsets[first_tl_row - day_row] += 1

            pt_hits, pt_row, pt_stride = find_pay_type_header(ws, day_row, day_cols)
            if pt_hits:
                order = tuple(tok for (_off, tok) in pt_hits)
                pay_type_order[order] += 1
                pay_type_header_offset[pt_row - day_row] += 1

            # Build a structural signature for this sheet
            sig = (
                ident[2] if ident else "no-identity",
                day_row,
                col_letter(first_col),
                stride,
                style,
                col_letter(tl[0]) if tl else None,
                tuple(tok for (_off, tok) in pt_hits) if pt_hits else None,
                (wc[0], col_letter(wc[1])) if wc else None,
                (st[0], col_letter(st[1])) if st else None,
            )
            profile_signatures[sig] += 1
            family = (
                "standard"  if (day_row == 4 and col_letter(first_col) == "L" and style == "short" and stride == 6) else
                "shifted"   if (day_row == 4 and col_letter(first_col) == "M" and style == "short" and stride == 6) else
                "fieldmech" if (day_row == 3 and col_letter(first_col) == "N" and style == "full"  and stride == 5) else
                "other"
            )
            layout_family[family] += 1
            file_details.append((f.name, sn, family, sig))

    print("=" * 70)
    print("STRUCTURAL CENSUS — input corpus")
    print("=" * 70)
    print(f"Workbooks scanned:                    {total_workbooks}")
    print(f"Total sheets across all workbooks:    {total_sheets_seen}")
    print(f"Reference / template / training:      {total_reference_by_name}")
    print(f"Employee-like sheets (counted below): {total_employee_like}")
    print()

    print("--- Identity layout (where name + EE ID live) ---")
    for k, n in identity_layouts.most_common():
        print(f"  {k:<20} {n}")
    print()

    print("--- Day-token style (MON vs MONDAY) ---")
    for k, n in day_token_style.most_common():
        print(f"  {k:<20} {n}")
    print()

    print("--- Day-label row position ---")
    for k, n in sorted(day_row_positions.items()):
        print(f"  row {k:<3}              {n}")
    print()

    print("--- First-day-column position ---")
    for k, n in day_first_col_positions.most_common():
        print(f"  col {k:<5}            {n}")
    print()

    print("--- Day-block stride (gap between consecutive day columns) ---")
    for k, n in sorted(day_stride.items(), key=lambda kv: (kv[0] is None, kv[0])):
        print(f"  stride {k!s:<5}         {n}")
    print()

    print("--- Date row offset from day-label row ---")
    for k, n in sorted(date_row_offsets.items()):
        print(f"  day_row + {k:<3}        {n}")
    print()

    print("--- Time-label column ---")
    for k, n in time_label_cols.most_common():
        print(f"  col {k:<5}            {n}")
    print()

    print("--- First time-label row offset from day-label row ---")
    for k, n in sorted(time_first_row_offsets.items()):
        print(f"  day_row + {k:<3}        {n}")
    print()

    print("--- Pay-type column order in day-block header ---")
    for k, n in pay_type_order.most_common():
        print(f"  {' / '.join(k):<35} {n}")
    print()

    print("--- WC State header position (row, col) ---")
    for (r, c), n in wc_positions.most_common():
        print(f"  {c}{r:<3}                {n}")
    print()

    print("--- ST header position (row, col) ---")
    for (r, c), n in st_positions.most_common():
        print(f"  {c}{r:<3}                {n}")
    print()

    print("--- Layout family (grouped) ---")
    for k, n in layout_family.most_common():
        print(f"  {k:<25} {n}")
    print()

    print(f"--- Distinct full structural signatures: {len(profile_signatures)} ---")
    for sig, n in profile_signatures.most_common():
        print(f"  ({n:>3})  {sig}")
    print()

    # Show which workbooks have which family
    print("--- Workbooks containing each family ---")
    family_to_files = defaultdict(set)
    for fname, sn, family, sig in file_details:
        family_to_files[family].add(fname)
    for family, fs in sorted(family_to_files.items()):
        print(f"  {family}: {len(fs)} workbook(s)")
        for fn in sorted(fs)[:5]:
            print(f"    - {fn}")
        if len(fs) > 5:
            print(f"    ... and {len(fs) - 5} more")
    print()


if __name__ == "__main__":
    census()
